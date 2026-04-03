import os
import asyncio
import json
import base64
import uuid
import time
import subprocess
import re
import html
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from flow import (
    create_requirement_analysis_flow,
    create_script_flow,
    create_system_test_flow,
    create_test_case_flow,
    create_test_design_flow,
)
from schemas import DEFAULT_ACTION_VOCABULARY, DEFAULT_ASSERTION_VOCABULARY, DEFAULT_CAPABILITIES
from utils.llm_client import LLMClient
from utils.rag_store import SimpleRAGStore


class GenerateRequest(BaseModel):
    requirements: List[str] = Field(..., min_length=1)
    rag_context: str = ""
    framework_capability_catalog: str = ""
    product_profile: Optional[Dict[str, Any]] = None
    test_environments: Optional[Dict[str, Any]] = None
    capabilities: Optional[Dict[str, Any]] = None
    action_vocabulary: Optional[List[str]] = None
    assertion_vocabulary: Optional[List[str]] = None


class PromptSettingsRequest(BaseModel):
    prompts: Dict[str, str]


class ExportTestCasesRequest(BaseModel):
    requirement_input: Optional[List[str]] = None
    requirement_spec: Optional[Dict[str, Any]] = None
    persona_reviews: Optional[Dict[str, Any]] = None
    requirement_review_history: Optional[Dict[str, Any]] = None
    test_design_spec: Optional[Dict[str, Any]] = None
    test_case_spec: Dict[str, Any]


class RAGUploadBase64Request(BaseModel):
    filename: str
    content_base64: str


class ChatTurn(BaseModel):
    role: str
    content: str


class ChatRequest(BaseModel):
    message: str
    history: List[ChatTurn] = Field(default_factory=list)


class AgenticCodingChatRequest(BaseModel):
    message: str
    history: List[ChatTurn] = Field(default_factory=list)
    context: Dict[str, Any] = Field(default_factory=dict)


class RequirementRoundRequest(BaseModel):
    requirements: List[str] = Field(..., min_length=1)
    rag_context: str = ""
    framework_capability_catalog: str = ""
    product_profile: Optional[Dict[str, Any]] = None
    test_environments: Optional[Dict[str, Any]] = None
    round: int = 1
    previous_persona_reviews: Optional[Dict[str, Any]] = None
    open_question_answers: Optional[Dict[str, str]] = None


class RequirementHITLStartRequest(BaseModel):
    requirements: List[str] = Field(..., min_length=1)
    rag_context: str = ""
    framework_capability_catalog: str = ""
    product_profile: Optional[Dict[str, Any]] = None
    test_environments: Optional[Dict[str, Any]] = None
    automation_modes: Optional[List[str]] = None
    adb_device_id: str = ""
    at_port: str = ""
    at_baudrate: int = 115200


class RequirementHITLNextRequest(BaseModel):
    open_question_answers: Optional[Dict[str, str]] = None


class StageGenerateRequest(BaseModel):
    requirement_input: Optional[List[str]] = None
    requirement_spec: Dict[str, Any]
    persona_reviews: Optional[Dict[str, Any]] = None
    requirement_review_history: Optional[Dict[str, Any]] = None
    test_design_spec: Optional[Dict[str, Any]] = None
    test_case_spec: Optional[Dict[str, Any]] = None
    product_profile: Optional[Dict[str, Any]] = None
    test_environments: Optional[Dict[str, Any]] = None
    capabilities: Optional[Dict[str, Any]] = None
    action_vocabulary: Optional[List[str]] = None
    assertion_vocabulary: Optional[List[str]] = None
    history_record_id: str = ""


class AutomationRequest(BaseModel):
    mode: str = Field(..., description="android_adb | at_serial")
    device_id: str = ""
    at_port: str = ""
    baudrate: int = 115200
    test_case_spec: Optional[Dict[str, Any]] = None
    script_spec: Optional[Dict[str, Any]] = None
    max_cases: int = 5
    max_steps: int = 15


class AtAssetSaveRequest(BaseModel):
    data: Dict[str, Any]
    locked_baseline: bool = False


class AtCompileRequest(BaseModel):
    use_llm: bool = True


class AtConfigCompileRequest(BaseModel):
    request_text: str = Field(..., min_length=1)
    use_llm: bool = True
    apply_changes: bool = True
    compile_after_apply: bool = True


app = FastAPI(title="PocketFlow System Test Agent Web")
base_dir = os.path.dirname(__file__)
static_dir = os.path.join(base_dir, "static")
prompt_dir = Path(base_dir) / "prompts"
at_agent_dir = Path(base_dir) / "at_agent"
at_specs_dir = at_agent_dir / "specs"
at_profiles_dir = at_agent_dir / "profiles"
at_models_dir = at_agent_dir / "models"
at_manifests_dir = at_agent_dir / "manifests"
at_extensions_dir = at_agent_dir / "extensions"
at_build_dir = at_agent_dir / "build"
at_spec_path = at_specs_dir / "3gpp_base_atspec.v0.json"
at_profile_path = at_profiles_dir / "generic_3gpp.profile.v0.json"
at_efsm_path = at_models_dir / "3gpp_base.efsm.json"
at_manifest_path = at_manifests_dir / "default.manifest.json"
at_extension_path = at_extensions_dir / "vendor.extension.json"
at_effective_atspec_path = at_build_dir / "effective_atspec.json"
at_effective_profile_path = at_build_dir / "effective_profile.json"
at_active_efsm_path = at_build_dir / "active_efsm.json"
at_compile_report_path = at_build_dir / "compile_report.json"
design_history_dir = Path(base_dir) / "history" / "designs"
rag_store = SimpleRAGStore(Path(base_dir) / "rag_data")
requirement_hitl_jobs: Dict[str, Dict[str, Any]] = {}
stage_jobs: Dict[str, Dict[str, Any]] = {}
ALLOWED_PROMPT_FILES = {
    "req_parse.txt",
    "review_spec_lawyer.txt",
    "review_carrier.txt",
    "review_ux_advocate.txt",
    "synthesis_arbiter.txt",
    "test_designer.txt",
    "integrated_matrix_designer.txt",
    "test_case_generator.txt",
    "harness_mapper.txt",
    "script_writer.txt",
}
app.mount("/static", StaticFiles(directory=static_dir), name="static")


ATSPEC_DEFAULT: Dict[str, Any] = {
    "meta": {
        "id": "3gpp.base",
        "version": "0.1.0",
        "sources": [
            "3GPP TS 27.007 Rel-18 (ETSI TS 127 007 V18.6.0)",
            "3GPP TS 27.005 Rel-18 (ETSI TS 127 005 V18.0.0)",
        ],
    },
    "result_codes": {
        "final_ok": ["OK"],
        "final_error": ["ERROR"],
        "final_cme_error_prefix": "+CME ERROR:",
        "final_cms_error_prefix": "+CMS ERROR:",
    },
    "interaction": {
        "line_ending_tx": "\\r",
        "line_split_rx": ["\\r\\n", "\\n"],
        "prompt": {"chars": [">"], "sms_end_ctrlz": 26, "sms_cancel_esc": 27},
    },
    "capabilities": [
        {"id": "te_ta.formatting", "desc": "E/verbose errors", "depends": [], "signals": []},
        {"id": "device.identity", "desc": "CGMI/CGMM/CGMR/CGSN", "depends": ["te_ta.formatting"], "signals": []},
        {"id": "sim.pin", "desc": "CPIN", "depends": ["te_ta.formatting"], "signals": []},
        {"id": "phone.functionality", "desc": "CFUN", "depends": ["sim.pin"], "signals": []},
        {"id": "net.registration.cs", "desc": "CREG", "depends": ["phone.functionality"], "signals": ["+CREG:"]},
        {"id": "ps.attach", "desc": "CGATT", "depends": ["phone.functionality"], "signals": []},
        {"id": "pdp.define", "desc": "CGDCONT", "depends": ["ps.attach"], "signals": []},
        {"id": "pdp.activate", "desc": "CGACT", "depends": ["pdp.define"], "signals": []},
        {"id": "sms.core", "desc": "SMS core", "depends": ["te_ta.formatting", "sim.pin"], "signals": ["+CMTI:", "+CMT:", "+CDS:"]},
    ],
    "commands": [
        {"id": "v250.ate", "capability": "te_ta.formatting", "at": "ATE{val}", "ops": ["set"]},
        {"id": "v250.atv", "capability": "te_ta.formatting", "at": "ATV{val}", "ops": ["set"]},
        {"id": "cmd.cmee", "capability": "te_ta.formatting", "at": "AT+CMEE={n}", "ops": ["set", "read", "test"]},
        {"id": "cmd.cgmi", "capability": "device.identity", "at": "AT+CGMI", "ops": ["action"]},
        {"id": "cmd.cgmm", "capability": "device.identity", "at": "AT+CGMM", "ops": ["action"]},
        {"id": "cmd.cgmr", "capability": "device.identity", "at": "AT+CGMR", "ops": ["action"]},
        {"id": "cmd.cgsn", "capability": "device.identity", "at": "AT+CGSN", "ops": ["action"]},
        {"id": "cmd.cpin", "capability": "sim.pin", "at": "AT+CPIN{arg}", "ops": ["read", "set"]},
        {"id": "cmd.cfun", "capability": "phone.functionality", "at": "AT+CFUN={fun}", "ops": ["set", "read"]},
        {"id": "cmd.creg", "capability": "net.registration.cs", "at": "AT+CREG={n}", "ops": ["set", "read"]},
        {"id": "cmd.cgatt", "capability": "ps.attach", "at": "AT+CGATT={state}", "ops": ["set", "read"]},
        {"id": "cmd.cgdcont", "capability": "pdp.define", "at": "AT+CGDCONT={cid},\"{pdp_type}\",\"{apn}\"", "ops": ["set", "read"]},
        {"id": "cmd.cgact", "capability": "pdp.activate", "at": "AT+CGACT={state},{cid}", "ops": ["set", "read"]},
        {"id": "sms.cmgf", "capability": "sms.core", "at": "AT+CMGF={mode}", "ops": ["set", "read"]},
        {"id": "sms.cnmi", "capability": "sms.core", "at": "AT+CNMI={mode},{mt},{bm},{ds},{bfr}", "ops": ["set", "read"]},
        {"id": "sms.cmgs", "capability": "sms.core", "at": "AT+CMGS={da}", "ops": ["action"], "expect": {"prompt": ">"}},
    ],
}

ATPROFILE_DEFAULT: Dict[str, Any] = {
    "meta": {"id": "profile.generic_3gpp", "version": "0.1.0"},
    "transport": {"baudrate": 115200, "data_bits": 8, "parity": "N", "stop_bits": 1, "read_encoding": "latin-1", "line_ending_tx": "\r"},
    "defaults": {
        "init_sequence": [
            {"cmd_id": "v250.ate", "params": {"val": 0}},
            {"cmd_id": "v250.atv", "params": {"val": 1}},
            {"cmd_id": "cmd.cmee", "params": {"n": 2}},
        ],
        "timeouts": {"default_sec": 3, "network_register_sec": 180, "pdp_activate_sec": 60, "sms_send_sec": 60},
    },
    "bindings": [
        {"capability": "device.identity", "impl": [{"cmd_id": "cmd.cgmi"}, {"cmd_id": "cmd.cgmm"}, {"cmd_id": "cmd.cgmr"}, {"cmd_id": "cmd.cgsn"}]},
        {"capability": "sim.pin", "impl": [{"cmd_id": "cmd.cpin"}]},
        {"capability": "phone.functionality", "impl": [{"cmd_id": "cmd.cfun"}]},
        {"capability": "net.registration.cs", "impl": [{"cmd_id": "cmd.creg"}]},
        {"capability": "ps.attach", "impl": [{"cmd_id": "cmd.cgatt"}]},
        {"capability": "pdp.define", "impl": [{"cmd_id": "cmd.cgdcont"}]},
        {"capability": "pdp.activate", "impl": [{"cmd_id": "cmd.cgact"}]},
        {"capability": "sms.core", "impl": [{"cmd_id": "sms.cmgf"}, {"cmd_id": "sms.cnmi"}, {"cmd_id": "sms.cmgs"}]},
    ],
    "vendor_overrides": {"example_quectel": {"add_cmds": [], "replace_bindings": []}},
}

ATEFSM_DEFAULT: Dict[str, Any] = {
    "meta": {"id": "efsm.3gpp_base", "version": "0.1.0"},
    "states": [
        {"id": "S0_BOOT"}, {"id": "S1_AT_READY"}, {"id": "S2_SIM_LOCKED"}, {"id": "S3_SIM_READY"},
        {"id": "S4_RF_ON"}, {"id": "S5_CS_REGISTERED"}, {"id": "S6_PS_ATTACHED"},
        {"id": "S7_PDP_DEFINED"}, {"id": "S8_PDP_ACTIVE"}, {"id": "S9_SMS_READY"},
    ],
    "transitions": [
        {"id": "T_INIT", "from": "S0_BOOT", "to": "S1_AT_READY", "action": {"cmd_sequence": ["v250.ate", "v250.atv", "cmd.cmee"]}, "coverage": {"points": ["init.sequence"]}},
        {"id": "T_CHECK_SIM", "from": "S1_AT_READY", "to": "S2_SIM_LOCKED", "action": {"cmd_id": "cmd.cpin", "params": {"arg": ""}}, "coverage": {"points": ["sim.pin.required"]}},
        {"id": "T_UNLOCK_SIM", "from": "S2_SIM_LOCKED", "to": "S3_SIM_READY", "action": {"cmd_id": "cmd.cpin", "params": {"arg": "=\"{PIN}\""}}, "coverage": {"points": ["sim.pin.unlock"]}},
        {"id": "T_RF_ON", "from": "S3_SIM_READY", "to": "S4_RF_ON", "action": {"cmd_id": "cmd.cfun", "params": {"fun": 1}}, "coverage": {"points": ["rf.on"]}},
        {"id": "T_ATTACH", "from": "S4_RF_ON", "to": "S6_PS_ATTACHED", "action": {"cmd_id": "cmd.cgatt", "params": {"state": 1}}, "coverage": {"points": ["ps.attach"]}},
        {"id": "T_DEFINE_PDP", "from": "S6_PS_ATTACHED", "to": "S7_PDP_DEFINED", "action": {"cmd_id": "cmd.cgdcont", "params": {"cid": 1, "pdp_type": "IPV4V6", "apn": "{APN}"}}, "coverage": {"points": ["pdp.define.cid1"]}},
        {"id": "T_ACTIVATE_PDP", "from": "S7_PDP_DEFINED", "to": "S8_PDP_ACTIVE", "action": {"cmd_id": "cmd.cgact", "params": {"state": 1, "cid": 1}}, "coverage": {"points": ["pdp.activate.cid1"]}},
        {"id": "T_SMS_READY", "from": "S3_SIM_READY", "to": "S9_SMS_READY", "action": {"cmd_sequence": ["sms.cmgf", "sms.cnmi"]}, "coverage": {"points": ["sms.setup"]}},
    ],
}

ATMANIFEST_DEFAULT: Dict[str, Any] = {
    "baseline": "atspec.3gpp@0.2",
    "extensions": ["atspec.vendor.custom@1.0"],
    "policy": {
        "must_have_capabilities": [
            "device.functional_level",
            "net.registration.cs",
            "ps.attach",
            "pdp.define",
            "pdp.activate",
            "sms.basic",
            "cs.call.basic",
        ],
        "allowed_missing_capabilities": [],
    },
    "test_scope": {
        "enable_capabilities": ["*"],
        "disable_capabilities": [],
        "enable_commands": [],
        "disable_commands": [],
    },
    "env": {
        "apn": "internet",
        "pin_secret_ref": "vault://sim_pin",
        "mo_call_number": "+8210xxxxxxx",
        "sms_da": "+8210yyyyyyy",
    },
}

ATEXTENSION_DEFAULT: Dict[str, Any] = {
    "meta": {"id": "atspec.vendor.custom", "version": "1.0"},
    "capabilities": [],
    "commands": [],
}


def _ensure_at_agent_assets():
    at_specs_dir.mkdir(parents=True, exist_ok=True)
    at_profiles_dir.mkdir(parents=True, exist_ok=True)
    at_models_dir.mkdir(parents=True, exist_ok=True)
    at_manifests_dir.mkdir(parents=True, exist_ok=True)
    at_extensions_dir.mkdir(parents=True, exist_ok=True)
    at_build_dir.mkdir(parents=True, exist_ok=True)
    if not at_spec_path.exists():
        at_spec_path.write_text(json.dumps(ATSPEC_DEFAULT, ensure_ascii=False, indent=2), encoding="utf-8")
    if not at_profile_path.exists():
        at_profile_path.write_text(json.dumps(ATPROFILE_DEFAULT, ensure_ascii=False, indent=2), encoding="utf-8")
    if not at_efsm_path.exists():
        at_efsm_path.write_text(json.dumps(ATEFSM_DEFAULT, ensure_ascii=False, indent=2), encoding="utf-8")
    if not at_manifest_path.exists():
        at_manifest_path.write_text(json.dumps(ATMANIFEST_DEFAULT, ensure_ascii=False, indent=2), encoding="utf-8")
    if not at_extension_path.exists():
        at_extension_path.write_text(json.dumps(ATEXTENSION_DEFAULT, ensure_ascii=False, indent=2), encoding="utf-8")


_ensure_at_agent_assets()
design_history_dir.mkdir(parents=True, exist_ok=True)


def _valid_history_record_id(record_id: str) -> bool:
    return bool(re.fullmatch(r"[A-Za-z0-9_-]{6,64}", record_id or ""))


def _history_record_dir(record_id: str) -> Path:
    if not _valid_history_record_id(record_id):
        raise HTTPException(status_code=404, detail="history记录不存在")
    return design_history_dir / record_id


def _history_record_json_path(record_id: str) -> Path:
    return _history_record_dir(record_id) / "record.json"


def _history_document_path(record_id: str) -> Path:
    return _history_record_dir(record_id) / "design_document.html"


def _history_document_url(record_id: str) -> str:
    return f"/api/design-history/{record_id}/document"


def _safe_json_text(value: Any) -> str:
    try:
        return json.dumps(value, ensure_ascii=False, indent=2)
    except Exception:
        return str(value)


def _history_title(requirement_input: List[str], requirement_spec: Dict[str, Any]) -> str:
    reqs = requirement_spec.get("final_requirements", []) if isinstance(requirement_spec, dict) else []
    if isinstance(reqs, list):
        for item in reqs:
            if isinstance(item, dict) and str(item.get("title", "")).strip():
                return str(item.get("title")).strip()
    for raw in requirement_input or []:
        text = str(raw or "").strip()
        if text:
            return text[:80]
    return "未命名测试设计"


def _history_meta_counts(record: Dict[str, Any]) -> Dict[str, int]:
    requirement_spec = record.get("requirement_spec", {}) if isinstance(record, dict) else {}
    test_design_spec = record.get("test_design_spec", {}) if isinstance(record, dict) else {}
    test_case_spec = record.get("test_case_spec", {}) if isinstance(record, dict) else {}
    return {
        "requirements": len(requirement_spec.get("final_requirements", []) if isinstance(requirement_spec, dict) else []),
        "objectives": len(test_design_spec.get("objectives", []) if isinstance(test_design_spec, dict) else []),
        "matrix_rows": len(test_design_spec.get("integrated_matrix", []) if isinstance(test_design_spec, dict) else []),
        "testcases": len(test_case_spec.get("testcases", []) if isinstance(test_case_spec, dict) else []),
    }


def _history_summary(record: Dict[str, Any]) -> Dict[str, Any]:
    counts = _history_meta_counts(record)
    record_id = str(record.get("record_id", ""))
    return {
        "record_id": record_id,
        "title": record.get("title", "未命名测试设计"),
        "status": record.get("status", "design_completed"),
        "created_at": record.get("created_at", ""),
        "updated_at": record.get("updated_at", ""),
        "requirement_count": counts["requirements"],
        "objective_count": counts["objectives"],
        "matrix_row_count": counts["matrix_rows"],
        "testcase_count": counts["testcases"],
        "has_testcases": counts["testcases"] > 0,
        "document_url": _history_document_url(record_id),
    }


def _render_html_list(items: Any) -> str:
    arr = [str(x).strip() for x in (items or []) if str(x).strip()]
    if not arr:
        return "<p>无</p>"
    return "<ul>" + "".join(f"<li>{html.escape(x)}</li>" for x in arr) + "</ul>"


def _render_html_kv_table(rows: List[List[str]]) -> str:
    if not rows:
        return "<p>无</p>"
    body = "".join(
        f"<tr><th>{html.escape(str(k))}</th><td>{html.escape(str(v))}</td></tr>"
        for k, v in rows
    )
    return f"<table><tbody>{body}</tbody></table>"


def _render_test_environment_lines(test_environments: Dict[str, Any]) -> List[str]:
    if not isinstance(test_environments, dict):
        return []
    display = test_environments.get("display")
    if isinstance(display, list):
        lines = [str(x).strip() for x in display if str(x).strip()]
        if lines:
            return lines
    details = test_environments.get("selected_details")
    if isinstance(details, list):
        lines = []
        for item in details:
            if not isinstance(item, dict):
                continue
            label = str(item.get("label", "")).strip()
            desc = str(item.get("description", "")).strip()
            priority = str(item.get("priority", "")).strip()
            prefix = f"{priority}. " if priority else ""
            if label and desc:
                lines.append(f"{prefix}{label}: {desc}")
            elif label:
                lines.append(f"{prefix}{label}")
        if lines:
            return lines
    return []


def _format_test_environment_cell(env: Any) -> tuple[str, str]:
    if not isinstance(env, dict):
        return ("未指定", "")
    primary = str(env.get("primary_label") or env.get("primary") or "未指定").strip() or "未指定"
    alternatives = env.get("alternative_labels")
    if not isinstance(alternatives, list) or not alternatives:
        alternatives = env.get("alternatives") if isinstance(env.get("alternatives"), list) else []
    alt_text = "、".join(str(x).strip() for x in alternatives if str(x).strip())
    rationale = str(env.get("rationale", "")).strip()
    risk = str(env.get("availability_risk", "")).strip()

    parts = [primary]
    if alt_text:
        parts.append(f"备选: {alt_text}")
    if rationale:
        parts.append(f"依据: {rationale}")
    return ("；".join(parts), risk)


def _parse_answer_map(answer_map: Dict[str, Any], source_round: int) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    if not isinstance(answer_map, dict):
        return out
    for raw_key, raw_answer in answer_map.items():
        answer = str(raw_answer or "").strip()
        if not answer:
            continue
        key = str(raw_key or "")
        req_id, question = "", key
        if "::" in key:
            req_id, question = key.split("::", 1)
        out.append(
            {
                "source_round": source_round,
                "req_id": req_id.strip(),
                "question": question.strip(),
                "answer": answer,
            }
        )
    return out


def _merge_answered_questions(existing: Any, additions: Any) -> List[Dict[str, Any]]:
    merged: Dict[tuple[str, str], Dict[str, Any]] = {}
    for item in list(existing or []) + list(additions or []):
        if not isinstance(item, dict):
            continue
        req_id = str(item.get("req_id", "")).strip()
        question = str(item.get("question", "")).strip()
        answer = str(item.get("answer", "")).strip()
        if not question or not answer:
            continue
        merged[(req_id, question)] = {
            "source_round": int(item.get("source_round", 0) or 0),
            "req_id": req_id,
            "question": question,
            "answer": answer,
        }
    return list(merged.values())


def _build_requirement_review_history(rounds: Any, answered_questions: Any) -> Dict[str, Any]:
    round_list = [item for item in (rounds or []) if isinstance(item, dict)]
    answered_list = [item for item in (answered_questions or []) if isinstance(item, dict)]
    return {
        "rounds": round_list,
        "answered_questions": answered_list,
        "answered_count": len(answered_list),
    }


def _render_history_review_table(items: List[Dict[str, Any]]) -> str:
    if not items:
        return "<p>无</p>"
    rows = []
    for it in items:
        rows.append(
            "<tr>"
            f"<td>{html.escape(str(it.get('req_id', '')))}</td>"
            f"<td>{html.escape('；'.join(str(x) for x in it.get('issues', []) or []))}</td>"
            f"<td>{html.escape('；'.join(str(x) for x in it.get('open_questions', []) or []))}</td>"
            f"<td>{html.escape(str(it.get('rewrite_suggestion', '')))}</td>"
            "</tr>"
        )
    return (
        "<table><thead><tr><th>需求ID</th><th>问题项</th><th>待确认问题</th><th>改写建议</th></tr></thead>"
        f"<tbody>{''.join(rows)}</tbody></table>"
    )


def _render_answered_question_table(items: Any) -> str:
    rows = [it for it in (items or []) if isinstance(it, dict)]
    if not rows:
        return "<p>无</p>"
    body = []
    for item in rows:
        body.append(
            "<tr>"
            f"<td>{html.escape(str(item.get('source_round', '')))}</td>"
            f"<td>{html.escape(str(item.get('req_id', '')))}</td>"
            f"<td>{html.escape(str(item.get('question', '')))}</td>"
            f"<td>{html.escape(str(item.get('answer', '')))}</td>"
            "</tr>"
        )
    return (
        "<table><thead><tr><th>来源轮次</th><th>需求ID</th><th>已确认问题</th><th>用户确认/补充</th></tr></thead>"
        f"<tbody>{''.join(body)}</tbody></table>"
    )


def _render_design_document_html(record: Dict[str, Any]) -> str:
    requirement_input = record.get("requirement_input", []) or []
    product_profile = record.get("product_profile", {}) if isinstance(record.get("product_profile"), dict) else {}
    test_environments = record.get("test_environments", {}) if isinstance(record.get("test_environments"), dict) else {}
    requirement_spec = record.get("requirement_spec", {}) if isinstance(record.get("requirement_spec"), dict) else {}
    persona_reviews = record.get("persona_reviews", {}) if isinstance(record.get("persona_reviews"), dict) else {}
    requirement_review_history = record.get("requirement_review_history", {}) if isinstance(record.get("requirement_review_history"), dict) else {}
    test_design_spec = record.get("test_design_spec", {}) if isinstance(record.get("test_design_spec"), dict) else {}
    reqs = requirement_spec.get("final_requirements", []) if isinstance(requirement_spec.get("final_requirements"), list) else []
    objectives = test_design_spec.get("objectives", []) if isinstance(test_design_spec.get("objectives"), list) else []
    matrices = test_design_spec.get("coverage_matrices", []) if isinstance(test_design_spec.get("coverage_matrices"), list) else []
    integrated = test_design_spec.get("integrated_matrix", []) if isinstance(test_design_spec.get("integrated_matrix"), list) else []
    title = record.get("title") or _history_title(requirement_input, requirement_spec)
    updated_at = record.get("updated_at", "")

    req_rows = []
    for item in reqs:
        req_rows.append(
            "<tr>"
            f"<td>{html.escape(str(item.get('req_id', '')))}</td>"
            f"<td>{html.escape(str(item.get('title', '')))}</td>"
            f"<td>{html.escape(str(item.get('priority', '')))}</td>"
            f"<td>{html.escape('、'.join(str(x) for x in item.get('rat_scope', []) or []))}</td>"
            f"<td>{html.escape(' / '.join(str(x) for x in ((item.get('acceptance') or {}).get('pass_fail', []) or [])))}</td>"
            "</tr>"
        )
    req_table = (
        "<table><thead><tr><th>需求ID</th><th>标题</th><th>优先级</th><th>RAT范围</th><th>通过标准</th></tr></thead>"
        f"<tbody>{''.join(req_rows)}</tbody></table>"
        if req_rows
        else "<p>无</p>"
    )

    review_sections = []
    for label, key in [("规范评审", "spec"), ("运营商评审", "carrier"), ("现网体验评审", "ux")]:
        reviews = (persona_reviews.get(key) or {}).get("reviews", []) if isinstance(persona_reviews.get(key), dict) else []
        review_sections.append(f"<section><h3>{html.escape(label)}</h3>{_render_history_review_table(reviews)}</section>")

    obj_rows = []
    for item in objectives:
        obj_rows.append(
            "<tr>"
            f"<td>{html.escape(str(item.get('objective_id', '')))}</td>"
            f"<td>{html.escape('、'.join(str(x) for x in item.get('linked_reqs', []) or []))}</td>"
            f"<td>{html.escape(str(item.get('goal', '')))}</td>"
            f"<td>{html.escape(' / '.join(str(x) for x in item.get('success_criteria', []) or []))}</td>"
            f"<td>{html.escape(' / '.join(str(x) for x in item.get('evidence', []) or []))}</td>"
            f"<td>{html.escape(str(item.get('priority', '')))}</td>"
            "</tr>"
        )
    objective_table = (
        "<table><thead><tr><th>目标ID</th><th>关联需求</th><th>目标</th><th>成功标准</th><th>证据</th><th>优先级</th></tr></thead>"
        f"<tbody>{''.join(obj_rows)}</tbody></table>"
        if obj_rows
        else "<p>无</p>"
    )

    matrix_cards = []
    for matrix in matrices:
        dims = matrix.get("dimensions", []) if isinstance(matrix, dict) else []
        dim_html = "<br/>".join(
            f"{html.escape(str(d.get('name', '')))}: {html.escape('、'.join(str(x) for x in (d.get('values', []) or [])))}"
            for d in dims if isinstance(d, dict)
        )
        sampling = matrix.get("sampling_strategy", {}) if isinstance(matrix, dict) else {}
        matrix_cards.append(
            "<article class='matrix-card'>"
            f"<h4>{html.escape(str(matrix.get('matrix_id', '覆盖矩阵')))}</h4>"
            f"<div>{dim_html or '无'}</div>"
            f"<p><strong>采样策略：</strong>{html.escape(str(sampling.get('type', '')))}</p>"
            f"<p><strong>必选组合：</strong>{html.escape(_safe_json_text(sampling.get('must_include', [])))}</p>"
            f"<p><strong>排除组合：</strong>{html.escape(_safe_json_text(sampling.get('exclude', [])))}</p>"
            "</article>"
        )
    matrix_html = "".join(matrix_cards) if matrix_cards else "<p>无</p>"

    integrated_rows = []
    for idx, row in enumerate(integrated, start=1):
        cfg = row.get("key_configuration", {}) if isinstance(row, dict) and isinstance(row.get("key_configuration"), dict) else {}
        cfg_text = " / ".join(f"{k}: {v}" for k, v in cfg.items())
        env_text, risk_text = _format_test_environment_cell(row.get("test_environment"))
        integrated_rows.append(
            "<tr>"
            f"<td>{html.escape(str(row.get('row_id', f'ROW-{idx:03d}')))}</td>"
            f"<td>{html.escape(str(row.get('req_id', '')))}</td>"
            f"<td>{html.escape(str(row.get('objective_id', '')))}</td>"
            f"<td>{html.escape(str(row.get('scenario', '')))}</td>"
            f"<td>{html.escape(cfg_text)}</td>"
            f"<td>{html.escape(' / '.join(str(x) for x in row.get('pass_criteria', []) or []))}</td>"
            f"<td>{html.escape(env_text)}</td>"
            f"<td>{html.escape(risk_text)}</td>"
            "</tr>"
        )
    integrated_table = (
        "<table><thead><tr><th>行ID</th><th>需求ID</th><th>关联目标</th><th>测试场景</th><th>配置选取</th><th>通过标准</th><th>推荐测试环境</th><th>资源风险</th></tr></thead>"
        f"<tbody>{''.join(integrated_rows)}</tbody></table>"
        if integrated_rows
        else "<p>无</p>"
    )

    profile_lines = []
    for item in product_profile.get("profile_display", []) if isinstance(product_profile.get("profile_display"), list) else []:
        if isinstance(item, dict):
            group = str(item.get("group", "")).strip()
            values = item.get("values", []) or []
            if group and values:
                profile_lines.append(f"{group}: {'、'.join(str(x) for x in values)}")
    test_environment_lines = _render_test_environment_lines(test_environments)
    answered_question_table = _render_answered_question_table(requirement_review_history.get("answered_questions", []))

    return f"""<!DOCTYPE html>
<html lang="zh-CN">
  <head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>{html.escape(str(title))} - 测试设计文档</title>
    <style>
      body {{ margin: 0; padding: 32px; font-family: "Segoe UI", "PingFang SC", "Hiragino Sans GB", "Microsoft YaHei", sans-serif; background: #f4efe5; color: #1f2937; }}
      .page {{ max-width: 1180px; margin: 0 auto; background: #fffdf8; border: 1px solid #e5dccd; border-radius: 18px; box-shadow: 0 18px 48px rgba(15, 23, 42, 0.08); overflow: hidden; }}
      .hero {{ padding: 28px 32px 24px; background: linear-gradient(135deg, #f6efe3 0%, #fffaf1 58%, #eef6f7 100%); border-bottom: 1px solid #e9ddcb; }}
      .hero h1 {{ margin: 0; font-size: 30px; }}
      .hero p {{ margin: 8px 0 0; color: #5b6472; }}
      .section {{ padding: 24px 32px; border-top: 1px solid #efe6d8; }}
      .section:first-of-type {{ border-top: none; }}
      h2 {{ margin: 0 0 14px; font-size: 20px; }}
      h3 {{ margin: 0 0 10px; font-size: 16px; }}
      h4 {{ margin: 0 0 8px; font-size: 14px; }}
      table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
      th, td {{ border: 1px solid #e5dccd; padding: 9px 10px; text-align: left; vertical-align: top; }}
      th {{ background: #f8f1e7; }}
      ul {{ margin: 8px 0 0 18px; }}
      p {{ line-height: 1.7; margin: 8px 0; }}
      .grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }}
      .matrix-card {{ border: 1px solid #e7ddcf; border-radius: 12px; padding: 14px; background: #fff; margin-bottom: 12px; }}
      @media (max-width: 900px) {{ body {{ padding: 12px; }} .section, .hero {{ padding: 18px; }} .grid {{ grid-template-columns: 1fr; }} }}
    </style>
  </head>
  <body>
    <div class="page">
      <section class="hero">
        <h1>标准测试设计文档</h1>
        <p>{html.escape(str(title))}</p>
      </section>
      <section class="section">
        <h2>一、文档信息</h2>
        {_render_html_kv_table([["记录ID", record.get("record_id", "")], ["最近更新时间", updated_at], ["当前状态", record.get("status", "")]])}
      </section>
      <section class="section">
        <h2>二、需求输入与产品背景</h2>
        <div class="grid">
          <div><h3>原始需求输入</h3>{_render_html_list(requirement_input)}</div>
          <div><h3>产品通信背景与支持特性</h3>{_render_html_list(profile_lines)}</div>
          <div><h3>测试团队可用测试环境</h3>{_render_html_list(test_environment_lines)}</div>
        </div>
      </section>
      <section class="section"><h2>三、测试需求解读稿</h2>{req_table}</section>
      <section class="section"><h2>四、三人评审记录</h2>{''.join(review_sections)}<section><h3>用户已确认问题</h3>{answered_question_table}</section></section>
      <section class="section"><h2>五、测试目标</h2>{objective_table}</section>
      <section class="section"><h2>六、覆盖矩阵</h2>{matrix_html}</section>
      <section class="section"><h2>七、整合覆盖矩阵</h2>{integrated_table}</section>
      <section class="section"><h2>八、设计说明</h2>{_render_html_list(test_design_spec.get("design_notes", []))}</section>
      <section class="section">
        <h2>九、范围裁剪与待跟进项</h2>
        <div class="grid">
          <div><h3>去范围项</h3>{_render_html_list(test_design_spec.get("de_scoped", []))}</div>
          <div><h3>待确认问题</h3>{_render_html_list(requirement_spec.get("questions_to_ask", []))}</div>
        </div>
      </section>
    </div>
  </body>
</html>"""


def _write_design_history_record(record: Dict[str, Any]) -> Dict[str, Any]:
    record_id = str(record.get("record_id") or uuid.uuid4().hex[:12])
    now = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    record["record_id"] = record_id
    record["title"] = _history_title(record.get("requirement_input", []) or [], record.get("requirement_spec", {}) or {})
    record["created_at"] = record.get("created_at") or now
    record["updated_at"] = now
    record.setdefault("status", "design_completed")
    record_dir = _history_record_dir(record_id)
    record_dir.mkdir(parents=True, exist_ok=True)
    _history_record_json_path(record_id).write_text(_safe_json_text(record), encoding="utf-8")
    _history_document_path(record_id).write_text(_render_design_document_html(record), encoding="utf-8")
    return record


def _load_design_history_record(record_id: str) -> Dict[str, Any]:
    path = _history_record_json_path(record_id)
    if not path.exists():
        raise HTTPException(status_code=404, detail="history记录不存在")
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"history记录损坏: {exc}")


def _upsert_design_history_record(
    *,
    record_id: str = "",
    requirement_input: Optional[List[str]] = None,
    requirement_spec: Optional[Dict[str, Any]] = None,
    persona_reviews: Optional[Dict[str, Any]] = None,
    requirement_review_history: Optional[Dict[str, Any]] = None,
    test_design_spec: Optional[Dict[str, Any]] = None,
    product_profile: Optional[Dict[str, Any]] = None,
    test_environments: Optional[Dict[str, Any]] = None,
    test_case_spec: Optional[Dict[str, Any]] = None,
    warnings: Optional[Dict[str, Any]] = None,
    trace: Optional[Dict[str, Any]] = None,
    status: Optional[str] = None,
) -> Dict[str, Any]:
    record: Dict[str, Any] = {}
    if record_id:
        try:
            record = _load_design_history_record(record_id)
        except HTTPException:
            record = {"record_id": record_id}
    record["requirement_input"] = requirement_input or record.get("requirement_input", []) or []
    record["requirement_spec"] = requirement_spec or record.get("requirement_spec", {}) or {}
    record["persona_reviews"] = persona_reviews or record.get("persona_reviews", {}) or {}
    record["requirement_review_history"] = requirement_review_history or record.get("requirement_review_history", {}) or {}
    record["test_design_spec"] = test_design_spec or record.get("test_design_spec", {}) or {}
    record["product_profile"] = product_profile or record.get("product_profile", {}) or {}
    record["test_environments"] = test_environments or record.get("test_environments", {}) or {}
    if test_case_spec is not None:
        record["test_case_spec"] = test_case_spec
    else:
        record.setdefault("test_case_spec", {})
    if warnings:
        merged = record.get("warnings", {}) if isinstance(record.get("warnings"), dict) else {}
        merged.update(warnings)
        record["warnings"] = merged
    else:
        record.setdefault("warnings", {})
    if trace:
        merged = record.get("trace", {}) if isinstance(record.get("trace"), dict) else {}
        merged.update(trace)
        record["trace"] = merged
    else:
        record.setdefault("trace", {})
    if status:
        record["status"] = status
    return _write_design_history_record(record)


def _list_design_history_records() -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    for path in sorted(design_history_dir.glob("*/record.json"), reverse=True):
        try:
            record = json.loads(path.read_text(encoding="utf-8"))
            items.append(_history_summary(record))
        except Exception:
            continue
    items.sort(key=lambda x: x.get("updated_at", ""), reverse=True)
    return items


@app.get("/")
def index():
    return FileResponse(os.path.join(static_dir, "index.html"))


@app.get("/api/design-history")
def list_design_history():
    return {"items": _list_design_history_records()}


@app.get("/api/design-history/{record_id}")
def get_design_history(record_id: str):
    record = _load_design_history_record(record_id)
    return {
        **record,
        "summary": _history_summary(record),
        "document_url": _history_document_url(record_id),
    }


@app.get("/api/design-history/{record_id}/document")
def get_design_history_document(record_id: str):
    record = _load_design_history_record(record_id)
    html_path = _history_document_path(record_id)
    if not html_path.exists():
        html_path.write_text(_render_design_document_html(record), encoding="utf-8")
    return FileResponse(html_path, media_type="text/html", filename=f"{record.get('title', record_id)}-测试设计文档.html")


def _build_effective_rag_context(requirements: List[str], manual_rag_context: str) -> tuple[str, List[Dict[str, Any]]]:
    query = "\n".join(requirements).strip()
    hits = rag_store.retrieve(query, top_k=6)
    auto_context = ""
    if hits:
        lines = []
        for i, h in enumerate(hits, start=1):
            text = (h.get("text", "") or "").strip()
            excerpt = text[:700]
            lines.append(
                f"[RAG-{i}] doc={h.get('filename')} chunk={h.get('chunk_id')} score={h.get('score')}\n{excerpt}"
            )
        auto_context = "\n\n".join(lines)

    parts = [manual_rag_context.strip(), auto_context.strip()]
    merged = "\n\n".join([p for p in parts if p])
    return merged, hits


@app.get("/api/rag/docs")
def rag_docs():
    return {"docs": rag_store.list_docs()}


@app.post("/api/rag/upload")
async def rag_upload(payload: RAGUploadBase64Request):
    filename = (payload.filename or "").strip()
    if not filename:
        raise HTTPException(status_code=400, detail="filename不能为空")
    try:
        content = base64.b64decode(payload.content_base64)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"base64解析失败: {exc}")
    if not content:
        raise HTTPException(status_code=400, detail="文件内容为空")
    try:
        meta = rag_store.ingest_file(filename, content)
        return {"uploaded": meta}
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"文档入库失败: {exc}")


@app.post("/api/rag/upload-base64")
async def rag_upload_base64(payload: RAGUploadBase64Request):
    # Backward-compatible alias. Both endpoints are JSON only; no multipart is used.
    return await rag_upload(payload)


@app.post("/api/chat")
async def chat(payload: ChatRequest):
    message = (payload.message or "").strip()
    if not message:
        raise HTTPException(status_code=400, detail="message不能为空")

    client = LLMClient()
    if not client.enabled:
        return {"reply": "LLM未配置。请先设置 LLM_API_BASE / LLM_API_KEY / LLM_MODEL。"}

    history_lines: List[str] = []
    for turn in (payload.history or [])[-8:]:
        role = "用户" if turn.role == "user" else "助手"
        txt = (turn.content or "").strip()
        if txt:
            history_lines.append(f"{role}: {txt}")
    history_text = "\n".join(history_lines)
    user_prompt = (
        (f"对话历史:\n{history_text}\n\n" if history_text else "")
        + f"用户最新问题:\n{message}\n\n请直接回答，简洁、准确、可执行。"
    )
    try:
        reply = client.generate(
            system_prompt="You are a helpful website chatbot. Be concise, accurate, and actionable.",
            user_prompt=user_prompt,
            temperature=0.3,
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"chat失败: {exc}")
    return {"reply": reply}


@app.post("/api/agentic-coding/chat")
async def agentic_coding_chat(payload: AgenticCodingChatRequest):
    message = (payload.message or "").strip()
    if not message:
        raise HTTPException(status_code=400, detail="message不能为空")

    client = LLMClient()
    if not client.enabled:
        return {"reply": "LLM未配置。请先设置 LLM_API_BASE / LLM_API_KEY / LLM_MODEL。"}

    history_lines: List[str] = []
    for turn in (payload.history or [])[-12:]:
        role = "用户" if turn.role == "user" else "助手"
        txt = (turn.content or "").strip()
        if txt:
            history_lines.append(f"{role}: {txt}")
    history_text = "\n".join(history_lines)

    ctx = payload.context or {}
    context_text = json.dumps(ctx, ensure_ascii=False, indent=2)
    user_prompt = (
        "你正在协助用户做测试脚本 Agentic Coding（生成、修改、debug）。\n"
        "请优先结合给定上下文中的最新脚本与联机调试结果。\n"
        "如果需要修改代码，请输出可直接替换的代码块（```python ... ```）。\n"
        "如果是定位问题，请给出最小改动建议和原因。\n\n"
        + (f"对话历史:\n{history_text}\n\n" if history_text else "")
        + f"上下文:\n{context_text}\n\n"
        + f"用户最新问题:\n{message}"
    )
    try:
        reply = client.generate(
            system_prompt="You are a senior test automation engineer for agentic coding. Be concrete and code-first.",
            user_prompt=user_prompt,
            temperature=0.2,
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"agentic coding chat失败: {exc}")
    return {"reply": reply}

@app.post("/api/generate")
async def generate(payload: GenerateRequest):
    reqs = [item.strip() for item in payload.requirements if item.strip()]
    if not reqs:
        raise HTTPException(status_code=400, detail="requirements不能为空")

    effective_rag_context, rag_hits = _build_effective_rag_context(reqs, payload.rag_context.strip())
    shared = {
        "input_requirements": reqs,
        "rag_context": effective_rag_context,
        "framework_capability_catalog": payload.framework_capability_catalog.strip(),
        "product_profile": payload.product_profile or {},
        "test_environments": payload.test_environments or {},
        "capabilities": payload.capabilities or DEFAULT_CAPABILITIES,
        "action_vocabulary": payload.action_vocabulary or DEFAULT_ACTION_VOCABULARY,
        "assertion_vocabulary": payload.assertion_vocabulary or DEFAULT_ASSERTION_VOCABULARY,
        "llm_client": LLMClient(),
        "warnings": [],
        "trace": [],
        "review_feedback": "",
    }

    flow = create_system_test_flow()
    try:
        await flow.run_async(shared)
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail={
                "error": str(exc),
                "failed_node": shared.get("current_node"),
                "trace": shared.get("trace", []),
                "warnings": shared.get("warnings", []),
            },
        )

    result = shared.get("final_result")

    if not result:
        raise HTTPException(status_code=500, detail="流程未产生结果")

    if isinstance(result, dict):
        result.setdefault("rag", {})
        result["rag"]["hits"] = rag_hits
        result["rag"]["context_used"] = bool(effective_rag_context)
    return result


def _collect_review_issues(persona_reviews: Dict[str, Any]) -> List[str]:
    issues: List[str] = []
    for persona in ["spec", "carrier", "ux"]:
        for item in (persona_reviews.get(persona, {}) or {}).get("reviews", []) or []:
            req_id = item.get("req_id", "")
            for issue in item.get("issues", []) or []:
                issues.append(f"[{persona}|{req_id}] {issue}")
    return issues


def _collect_open_questions(persona_reviews: Dict[str, Any]) -> List[Dict[str, str]]:
    out: List[Dict[str, str]] = []
    seen = set()
    for persona in ["spec", "carrier", "ux"]:
        for item in (persona_reviews.get(persona, {}) or {}).get("reviews", []) or []:
            req_id = item.get("req_id", "")
            for q in item.get("open_questions", []) or []:
                k = (req_id, q)
                if k in seen:
                    continue
                seen.add(k)
                out.append({"req_id": req_id, "question": q})
    return out


async def _emit_hitl_event(job: Dict[str, Any], ev: Dict[str, Any]):
    """Broadcast HITL events to internal control queue and all SSE subscribers."""
    control_q: asyncio.Queue = job["queue"]
    await control_q.put(ev)
    job.setdefault("recent_events", []).append(ev)
    if len(job["recent_events"]) > 80:
        job["recent_events"] = job["recent_events"][-80:]

    stale: List[asyncio.Queue] = []
    for sub_q in job.get("subscribers", []) or []:
        try:
            sub_q.put_nowait(ev)
        except Exception:
            stale.append(sub_q)
    if stale:
        job["subscribers"] = [q for q in (job.get("subscribers", []) or []) if q not in stale]


async def _emit_stage_event(job: Dict[str, Any], ev: Dict[str, Any]):
    """Broadcast stage events to all subscribers and keep a short replay window."""
    control_q: asyncio.Queue = job["queue"]
    await control_q.put(ev)
    job.setdefault("recent_events", []).append(ev)
    if len(job["recent_events"]) > 120:
        job["recent_events"] = job["recent_events"][-120:]

    stale: List[asyncio.Queue] = []
    for sub_q in job.get("subscribers", []) or []:
        try:
            sub_q.put_nowait(ev)
        except Exception:
            stale.append(sub_q)
    if stale:
        job["subscribers"] = [q for q in (job.get("subscribers", []) or []) if q not in stale]


async def _run_requirement_hitl_job(job_id: str):
    job = requirement_hitl_jobs[job_id]
    queue: asyncio.Queue = job["queue"]
    reqs: List[str] = job["requirements"]
    rag_context: str = job["rag_context"]
    framework_capability_catalog: str = job["framework_capability_catalog"]
    product_profile: Dict[str, Any] = job.get("product_profile") or {}
    test_environments: Dict[str, Any] = job.get("test_environments") or {}

    try:
        for round_no in range(1, 4):
            effective_rag_context, rag_hits = _build_effective_rag_context(reqs, rag_context)
            shared = {
                "input_requirements": reqs,
                "rag_context": effective_rag_context,
                "framework_capability_catalog": framework_capability_catalog,
                "product_profile": product_profile,
                "test_environments": test_environments,
                "capabilities": DEFAULT_CAPABILITIES,
                "action_vocabulary": DEFAULT_ACTION_VOCABULARY,
                "assertion_vocabulary": DEFAULT_ASSERTION_VOCABULARY,
                "llm_client": LLMClient(),
                "warnings": [],
                "trace": [],
                "review_feedback": job.get("review_feedback", ""),
                "event_queue": queue,
            }
            flow = create_requirement_analysis_flow()
            await flow.run_async(shared)

            requirement_spec = shared.get("requirement_spec", {})
            persona_reviews = shared.get("persona_reviews", {})
            round_summary = {
                "round": round_no,
                "is_final_round": round_no >= 3,
                "issues": _collect_review_issues(persona_reviews),
                "open_questions": _collect_open_questions(persona_reviews),
            }
            round_summaries = [item for item in (job.get("round_summaries", []) or []) if item.get("round") != round_no]
            round_summaries.append(round_summary)
            round_summaries.sort(key=lambda item: item.get("round", 0))
            job["round_summaries"] = round_summaries
            review_history = _build_requirement_review_history(job.get("round_summaries", []), job.get("answered_questions", []))
            round_payload = {
                "job_id": job_id,
                "round": round_no,
                "is_final_round": round_no >= 3,
                "requirement_spec": requirement_spec,
                "persona_reviews": persona_reviews,
                "issues": round_summary["issues"],
                "open_questions": round_summary["open_questions"],
                "requirement_review_history": review_history,
                "warnings": shared.get("warnings", []),
                "trace": shared.get("trace", []),
                "rag": {"hits": rag_hits, "context_used": bool(effective_rag_context)},
            }
            job["last_payload"] = round_payload
            job["round"] = round_no
            job["status"] = "round_ready"
            await _emit_hitl_event(job, {"type": "round_ready", "payload": round_payload})

            if round_no >= 3:
                job["status"] = "completed"
                await _emit_hitl_event(job, {"type": "completed", "payload": round_payload})
                return

            job["status"] = "waiting_answers"
            await _emit_hitl_event(
                job,
                {
                    "type": "waiting_answers",
                    "payload": {"job_id": job_id, "round": round_no, "open_questions": round_payload["open_questions"]},
                }
            )
            await job["answer_event"].wait()
            job["answer_event"].clear()
            answers = job.get("pending_answers", {}) or {}
            if answers:
                job["answered_questions"] = _merge_answered_questions(
                    job.get("answered_questions", []),
                    _parse_answer_map(answers, round_no),
                )

            fb_parts: List[str] = []
            issues = _collect_review_issues(persona_reviews)
            if issues:
                fb_parts.append("【上轮评审issues回灌】" + " | ".join(issues))
            answer_lines = [f"{k}: {v}" for k, v in answers.items() if str(v).strip()]
            if answer_lines:
                fb_parts.append("【用户补充回答】" + " | ".join(answer_lines))
            job["review_feedback"] = "\n".join(fb_parts).strip()
            job["pending_answers"] = {}

    except Exception as exc:
        job["status"] = "failed"
        job["error"] = str(exc)
        await _emit_hitl_event(job, {"type": "error", "payload": {"job_id": job_id, "error": str(exc)}})


async def _wait_hitl_round(job_id: str, timeout_s: int = 180):
    job = requirement_hitl_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job不存在")
    queue: asyncio.Queue = job["queue"]
    try:
        while True:
            ev = await asyncio.wait_for(queue.get(), timeout=timeout_s)
            et = ev.get("type")
            if et in {"round_ready", "completed"}:
                return ev.get("payload", {})
            if et == "error":
                msg = ev.get("payload", {}).get("error", "unknown error")
                raise HTTPException(status_code=500, detail=f"HITL执行失败: {msg}")
    except asyncio.TimeoutError:
        raise HTTPException(status_code=504, detail="等待HITL结果超时")


@app.post("/api/hitl/requirements/start")
async def hitl_requirements_start(payload: RequirementHITLStartRequest):
    reqs = [r.strip() for r in payload.requirements if r.strip()]
    if not reqs:
        raise HTTPException(status_code=400, detail="requirements不能为空")

    job_id = uuid.uuid4().hex[:12]
    queue: asyncio.Queue = asyncio.Queue()
    answer_event = asyncio.Event()
    modes = [m for m in (payload.automation_modes or []) if m]
    auto_lines = []
    if modes:
        auto_lines.append(f"Automation Modes: {', '.join(modes)}")
    if payload.adb_device_id.strip():
        auto_lines.append(f"ADB Device: {payload.adb_device_id.strip()}")
    if payload.at_port.strip():
        auto_lines.append(f"AT Port: {payload.at_port.strip()} @ {payload.at_baudrate}")
    framework_capability_catalog = "\n".join(
        [x for x in [payload.framework_capability_catalog.strip(), "\n".join(auto_lines).strip()] if x]
    )

    requirement_hitl_jobs[job_id] = {
        "job_id": job_id,
        "status": "starting",
        "round": 0,
        "requirements": reqs,
        "rag_context": payload.rag_context.strip(),
        "framework_capability_catalog": framework_capability_catalog,
        "product_profile": payload.product_profile or {},
        "test_environments": payload.test_environments or {},
        "review_feedback": "",
        "round_summaries": [],
        "answered_questions": [],
        "pending_answers": {},
        "answer_event": answer_event,
        "queue": queue,
        "subscribers": [],
        "recent_events": [],
        "last_payload": None,
        "error": "",
    }
    requirement_hitl_jobs[job_id]["task"] = asyncio.create_task(_run_requirement_hitl_job(job_id))
    first = await _wait_hitl_round(job_id)
    return first


@app.post("/api/hitl/requirements/{job_id}/next")
async def hitl_requirements_next(job_id: str, payload: RequirementHITLNextRequest):
    job = requirement_hitl_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job不存在")
    if job.get("status") not in {"waiting_answers", "round_ready"}:
        if job.get("status") == "completed":
            return job.get("last_payload") or {"job_id": job_id, "is_final_round": True}
        raise HTTPException(status_code=409, detail=f"当前状态不允许next: {job.get('status')}")
    job["pending_answers"] = payload.open_question_answers or {}
    job["answer_event"].set()
    nxt = await _wait_hitl_round(job_id)
    return nxt


@app.get("/api/hitl/requirements/{job_id}/state")
def hitl_requirements_state(job_id: str):
    job = requirement_hitl_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job不存在")
    return {
        "job_id": job_id,
        "status": job.get("status"),
        "round": job.get("round", 0),
        "last_payload": job.get("last_payload"),
        "error": job.get("error", ""),
    }


@app.get("/api/hitl/requirements/{job_id}/stream")
async def hitl_requirements_stream(job_id: str):
    job = requirement_hitl_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job不存在")

    sub_q: asyncio.Queue = asyncio.Queue()
    subscribers = job.setdefault("subscribers", [])
    subscribers.append(sub_q)

    async def event_generator():
        try:
            init_payload = {
                "job_id": job_id,
                "status": job.get("status"),
                "round": job.get("round", 0),
                "error": job.get("error", ""),
            }
            yield f"data: {json.dumps({'type': 'state', 'payload': init_payload}, ensure_ascii=False)}\n\n"

            for ev in job.get("recent_events", [])[-20:]:
                yield f"data: {json.dumps(ev, ensure_ascii=False)}\n\n"

            while True:
                ev = await sub_q.get()
                yield f"data: {json.dumps(ev, ensure_ascii=False)}\n\n"
                if ev.get("type") in {"completed", "error"}:
                    break
        finally:
            if sub_q in (job.get("subscribers", []) or []):
                job["subscribers"] = [q for q in job.get("subscribers", []) if q is not sub_q]

    return StreamingResponse(event_generator(), media_type="text/event-stream")


async def _run_stage_job(job_id: str):
    job = stage_jobs[job_id]
    stage = job["stage"]
    payload: StageGenerateRequest = job["payload"]
    queue: asyncio.Queue = job["queue"]

    await _emit_stage_event(job, {"type": "job_started", "payload": {"job_id": job_id, "stage": stage}})

    try:
        if stage == "design":
            shared = {
                "input_requirements": payload.requirement_input or [],
                "requirement_spec": payload.requirement_spec,
                "persona_reviews": payload.persona_reviews or {},
                "requirement_review_history": payload.requirement_review_history or {},
                "product_profile": payload.product_profile or {},
                "test_environments": payload.test_environments or {},
                "llm_client": LLMClient(),
                "warnings": [],
                "trace": [],
                "event_queue": queue,
            }
            flow = create_test_design_flow()
            await flow.run_async(shared)
            history_record = _upsert_design_history_record(
                record_id=payload.history_record_id or "",
                requirement_input=payload.requirement_input or [],
                requirement_spec=payload.requirement_spec or {},
                persona_reviews=payload.persona_reviews or {},
                requirement_review_history=payload.requirement_review_history or {},
                test_design_spec=shared.get("test_design_spec", {}),
                product_profile=payload.product_profile or {},
                test_environments=payload.test_environments or {},
                warnings={"design": shared.get("warnings", [])},
                trace={"design": shared.get("trace", [])},
                status="design_completed",
            )
            result = {
                "test_design_spec": shared.get("test_design_spec", {}),
                "warnings": shared.get("warnings", []),
                "trace": shared.get("trace", []),
                "design_history": _history_summary(history_record),
            }
        elif stage == "testcases":
            if not payload.test_design_spec:
                raise HTTPException(status_code=400, detail="缺少test_design_spec")
            shared = {
                "input_requirements": payload.requirement_input or [],
                "requirement_spec": payload.requirement_spec,
                "test_design_spec": payload.test_design_spec,
                "requirement_review_history": payload.requirement_review_history or {},
                "product_profile": payload.product_profile or {},
                "test_environments": payload.test_environments or {},
                "action_vocabulary": payload.action_vocabulary or DEFAULT_ACTION_VOCABULARY,
                "llm_client": LLMClient(),
                "warnings": [],
                "trace": [],
                "event_queue": queue,
            }
            flow = create_test_case_flow()
            await flow.run_async(shared)
            history_record = _upsert_design_history_record(
                record_id=payload.history_record_id or "",
                requirement_input=payload.requirement_input or [],
                requirement_spec=payload.requirement_spec or {},
                persona_reviews=payload.persona_reviews or {},
                requirement_review_history=payload.requirement_review_history or {},
                test_design_spec=payload.test_design_spec or {},
                product_profile=payload.product_profile or {},
                test_environments=payload.test_environments or {},
                test_case_spec=shared.get("test_case_spec", {}),
                warnings={"testcases": shared.get("warnings", [])},
                trace={"testcases": shared.get("trace", [])},
                status="testcases_completed",
            )
            result = {
                "test_case_spec": shared.get("test_case_spec", {}),
                "warnings": shared.get("warnings", []),
                "trace": shared.get("trace", []),
                "design_history": _history_summary(history_record),
            }
        elif stage == "scripts":
            if not payload.test_case_spec:
                raise HTTPException(status_code=400, detail="缺少test_case_spec")
            shared = {
                "test_case_spec": payload.test_case_spec,
                "product_profile": payload.product_profile or {},
                "test_environments": payload.test_environments or {},
                "capabilities": payload.capabilities or DEFAULT_CAPABILITIES,
                "assertion_vocabulary": payload.assertion_vocabulary or DEFAULT_ASSERTION_VOCABULARY,
                "llm_client": LLMClient(),
                "warnings": [],
                "trace": [],
                "event_queue": queue,
            }
            flow = create_script_flow()
            await flow.run_async(shared)
            result = {
                "script_spec": shared.get("script_spec", {}),
                "test_code_reference": shared.get("test_code_reference", ""),
                "warnings": shared.get("warnings", []),
                "trace": shared.get("trace", []),
            }
        else:
            raise RuntimeError(f"不支持的stage: {stage}")

        job["status"] = "completed"
        job["result"] = result
        await _emit_stage_event(
            job,
            {"type": "completed", "payload": {"job_id": job_id, "stage": stage, "result": result}},
        )
    except Exception as exc:
        job["status"] = "failed"
        detail = {
            "error": str(exc),
            "failed_node": (locals().get("shared") or {}).get("current_node"),
            "trace": (locals().get("shared") or {}).get("trace", []),
            "warnings": (locals().get("shared") or {}).get("warnings", []),
        }
        job["error"] = detail
        await _emit_stage_event(
            job,
            {"type": "error", "payload": {"job_id": job_id, "stage": stage, **detail}},
        )


@app.post("/api/stage/{stage}/start")
async def stage_start(stage: str, payload: StageGenerateRequest):
    stage_name = (stage or "").strip().lower()
    if stage_name not in {"design", "testcases", "scripts"}:
        raise HTTPException(status_code=400, detail="stage必须是design/testcases/scripts")

    if stage_name == "design" and not payload.requirement_spec:
        raise HTTPException(status_code=400, detail="缺少requirement_spec")
    if stage_name == "testcases" and not payload.test_design_spec:
        raise HTTPException(status_code=400, detail="缺少test_design_spec")
    if stage_name == "scripts" and not payload.test_case_spec:
        raise HTTPException(status_code=400, detail="缺少test_case_spec")

    job_id = uuid.uuid4().hex[:12]
    q: asyncio.Queue = asyncio.Queue()
    stage_jobs[job_id] = {
        "job_id": job_id,
        "stage": stage_name,
        "status": "running",
        "payload": payload,
        "queue": q,
        "subscribers": [],
        "recent_events": [],
        "result": None,
        "error": None,
    }
    stage_jobs[job_id]["task"] = asyncio.create_task(_run_stage_job(job_id))
    return {"job_id": job_id, "stage": stage_name, "status": "running"}


@app.get("/api/stage/{job_id}/state")
def stage_state(job_id: str):
    job = stage_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job不存在")
    return {
        "job_id": job_id,
        "stage": job.get("stage"),
        "status": job.get("status"),
        "result": job.get("result"),
        "error": job.get("error"),
    }


@app.get("/api/stage/{job_id}/stream")
async def stage_stream(job_id: str):
    job = stage_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job不存在")
    queue: asyncio.Queue = job["queue"]

    async def event_generator():
        init_payload = {
            "job_id": job_id,
            "stage": job.get("stage"),
            "status": job.get("status"),
            "error": job.get("error"),
        }
        yield f"data: {json.dumps({'type': 'state', 'payload': init_payload}, ensure_ascii=False)}\n\n"
        for ev in job.get("recent_events", [])[-40:]:
            yield f"data: {json.dumps(ev, ensure_ascii=False)}\n\n"

        while True:
            ev = await queue.get()
            yield f"data: {json.dumps(ev, ensure_ascii=False)}\n\n"
            if ev.get("type") in {"completed", "error"}:
                break

    return StreamingResponse(event_generator(), media_type="text/event-stream")


@app.post("/api/requirements/analyze")
async def requirements_analyze(payload: RequirementRoundRequest):
    round_no = max(1, min(3, payload.round))
    reqs = [r.strip() for r in payload.requirements if r.strip()]
    if not reqs:
        raise HTTPException(status_code=400, detail="requirements不能为空")

    review_feedback = ""
    if round_no > 1 and payload.previous_persona_reviews:
        issues = _collect_review_issues(payload.previous_persona_reviews)
        fb_parts: List[str] = []
        if issues:
            fb_parts.append("【上轮评审issues回灌】" + " | ".join(issues))
        if payload.open_question_answers:
            answers = [f"{k}: {v}" for k, v in payload.open_question_answers.items() if str(v).strip()]
            if answers:
                fb_parts.append("【用户补充回答】" + " | ".join(answers))
        review_feedback = "\n".join(fb_parts).strip()

    effective_rag_context, rag_hits = _build_effective_rag_context(reqs, payload.rag_context.strip())

    shared = {
        "input_requirements": reqs,
        "rag_context": effective_rag_context,
        "framework_capability_catalog": payload.framework_capability_catalog.strip(),
        "product_profile": payload.product_profile or {},
        "test_environments": payload.test_environments or {},
        "capabilities": DEFAULT_CAPABILITIES,
        "action_vocabulary": DEFAULT_ACTION_VOCABULARY,
        "assertion_vocabulary": DEFAULT_ASSERTION_VOCABULARY,
        "llm_client": LLMClient(),
        "warnings": [],
        "trace": [],
        "review_feedback": review_feedback,
    }

    flow = create_requirement_analysis_flow()
    try:
        await flow.run_async(shared)
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail={
                "error": str(exc),
                "failed_node": shared.get("current_node"),
                "trace": shared.get("trace", []),
                "warnings": shared.get("warnings", []),
            },
        )

    requirement_spec = shared.get("requirement_spec", {})
    persona_reviews = shared.get("persona_reviews", {})
    return {
        "round": round_no,
        "is_final_round": round_no >= 3,
        "requirement_spec": requirement_spec,
        "persona_reviews": persona_reviews,
        "issues": _collect_review_issues(persona_reviews),
        "open_questions": _collect_open_questions(persona_reviews),
        "warnings": shared.get("warnings", []),
        "trace": shared.get("trace", []),
        "rag": {"hits": rag_hits, "context_used": bool(effective_rag_context)},
    }


@app.post("/api/design/generate")
async def design_generate(payload: StageGenerateRequest):
    shared = {
        "input_requirements": payload.requirement_input or [],
        "requirement_spec": payload.requirement_spec,
        "persona_reviews": payload.persona_reviews or {},
        "requirement_review_history": payload.requirement_review_history or {},
        "product_profile": payload.product_profile or {},
        "test_environments": payload.test_environments or {},
        "llm_client": LLMClient(),
        "warnings": [],
        "trace": [],
    }
    flow = create_test_design_flow()
    try:
        await flow.run_async(shared)
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail={
                "error": str(exc),
                "failed_node": shared.get("current_node"),
                "trace": shared.get("trace", []),
                "warnings": shared.get("warnings", []),
            },
        )
    history_record = _upsert_design_history_record(
        record_id=payload.history_record_id or "",
        requirement_input=payload.requirement_input or [],
        requirement_spec=payload.requirement_spec or {},
        persona_reviews=payload.persona_reviews or {},
        requirement_review_history=payload.requirement_review_history or {},
        test_design_spec=shared.get("test_design_spec", {}),
        product_profile=payload.product_profile or {},
        test_environments=payload.test_environments or {},
        warnings={"design": shared.get("warnings", [])},
        trace={"design": shared.get("trace", [])},
        status="design_completed",
    )
    return {
        "test_design_spec": shared.get("test_design_spec", {}),
        "warnings": shared.get("warnings", []),
        "trace": shared.get("trace", []),
        "design_history": _history_summary(history_record),
    }


@app.post("/api/testcases/generate")
async def testcases_generate(payload: StageGenerateRequest):
    if not payload.test_design_spec:
        raise HTTPException(status_code=400, detail="缺少test_design_spec")
    shared = {
        "input_requirements": payload.requirement_input or [],
        "requirement_spec": payload.requirement_spec,
        "test_design_spec": payload.test_design_spec,
        "requirement_review_history": payload.requirement_review_history or {},
        "product_profile": payload.product_profile or {},
        "test_environments": payload.test_environments or {},
        "action_vocabulary": payload.action_vocabulary or DEFAULT_ACTION_VOCABULARY,
        "llm_client": LLMClient(),
        "warnings": [],
        "trace": [],
    }
    flow = create_test_case_flow()
    try:
        await flow.run_async(shared)
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail={
                "error": str(exc),
                "failed_node": shared.get("current_node"),
                "trace": shared.get("trace", []),
                "warnings": shared.get("warnings", []),
            },
        )
    history_record = _upsert_design_history_record(
        record_id=payload.history_record_id or "",
        requirement_input=payload.requirement_input or [],
        requirement_spec=payload.requirement_spec or {},
        persona_reviews=payload.persona_reviews or {},
        requirement_review_history=payload.requirement_review_history or {},
        test_design_spec=payload.test_design_spec or {},
        product_profile=payload.product_profile or {},
        test_environments=payload.test_environments or {},
        test_case_spec=shared.get("test_case_spec", {}),
        warnings={"testcases": shared.get("warnings", [])},
        trace={"testcases": shared.get("trace", [])},
        status="testcases_completed",
    )
    return {
        "test_case_spec": shared.get("test_case_spec", {}),
        "warnings": shared.get("warnings", []),
        "trace": shared.get("trace", []),
        "design_history": _history_summary(history_record),
    }


@app.post("/api/scripts/generate")
async def scripts_generate(payload: StageGenerateRequest):
    if not payload.test_case_spec:
        raise HTTPException(status_code=400, detail="缺少test_case_spec")
    shared = {
        "test_case_spec": payload.test_case_spec,
        "product_profile": payload.product_profile or {},
        "test_environments": payload.test_environments or {},
        "capabilities": payload.capabilities or DEFAULT_CAPABILITIES,
        "assertion_vocabulary": payload.assertion_vocabulary or DEFAULT_ASSERTION_VOCABULARY,
        "llm_client": LLMClient(),
        "warnings": [],
        "trace": [],
    }
    flow = create_script_flow()
    try:
        await flow.run_async(shared)
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail={
                "error": str(exc),
                "failed_node": shared.get("current_node"),
                "trace": shared.get("trace", []),
                "warnings": shared.get("warnings", []),
            },
        )
    return {
        "script_spec": shared.get("script_spec", {}),
        "test_code_reference": shared.get("test_code_reference", ""),
        "warnings": shared.get("warnings", []),
        "trace": shared.get("trace", []),
    }


@app.post("/api/generate/stream")
async def generate_stream(payload: GenerateRequest):
    reqs = [item.strip() for item in payload.requirements if item.strip()]
    if not reqs:
        raise HTTPException(status_code=400, detail="requirements不能为空")

    effective_rag_context, rag_hits = _build_effective_rag_context(reqs, payload.rag_context.strip())
    queue: asyncio.Queue = asyncio.Queue()
    shared = {
        "input_requirements": reqs,
        "rag_context": effective_rag_context,
        "framework_capability_catalog": payload.framework_capability_catalog.strip(),
        "product_profile": payload.product_profile or {},
        "test_environments": payload.test_environments or {},
        "capabilities": payload.capabilities or DEFAULT_CAPABILITIES,
        "action_vocabulary": payload.action_vocabulary or DEFAULT_ACTION_VOCABULARY,
        "assertion_vocabulary": payload.assertion_vocabulary or DEFAULT_ASSERTION_VOCABULARY,
        "llm_client": LLMClient(),
        "warnings": [],
        "trace": [],
        "event_queue": queue,
        "review_feedback": "",
    }

    flow = create_system_test_flow()

    async def runner():
        try:
            await flow.run_async(shared)
        except Exception as exc:
            await queue.put(
                {
                    "type": "error",
                    "payload": {
                        "error": str(exc),
                        "failed_node": shared.get("current_node"),
                        "trace": shared.get("trace", []),
                        "warnings": shared.get("warnings", []),
                    },
                }
            )
        finally:
            await queue.put({"type": "done", "payload": {}})

    async def event_generator():
        task = asyncio.create_task(runner())
        try:
            yield f"data: {json.dumps({'type': 'rag', 'payload': {'hits': rag_hits, 'context_used': bool(effective_rag_context)}}, ensure_ascii=False)}\n\n"
            while True:
                ev = await queue.get()
                yield f"data: {json.dumps(ev, ensure_ascii=False)}\n\n"
                if ev.get("type") == "done":
                    break
        finally:
            if not task.done():
                task.cancel()

    return StreamingResponse(event_generator(), media_type="text/event-stream")


@app.get("/api/settings/prompts")
def get_prompts():
    data: Dict[str, str] = {}
    for name in sorted(ALLOWED_PROMPT_FILES):
        p = prompt_dir / name
        if p.exists():
            data[name] = p.read_text(encoding="utf-8")
    return {"prompts": data}


def _flow_to_graph(flow_obj: Any) -> Dict[str, Any]:
    start = getattr(flow_obj, "start_node", None)
    if not start:
        return {"nodes": [], "links": [], "mermaid": "flowchart LR\n", "start_node_id": ""}

    node_id_map: Dict[int, str] = {}
    nodes: List[Dict[str, Any]] = []
    links: List[Dict[str, Any]] = []
    queue: List[Any] = [start]
    visited: set[int] = set()
    edge_seen: set[tuple[str, str, str]] = set()
    counter = 1

    while queue:
        node = queue.pop(0)
        obj_id = id(node)
        if obj_id in visited:
            continue
        visited.add(obj_id)
        if obj_id not in node_id_map:
            node_id_map[obj_id] = f"N{counter}"
            counter += 1
        nid = node_id_map[obj_id]
        nodes.append({"id": nid, "label": type(node).__name__})
        for action, nxt in (getattr(node, "successors", {}) or {}).items():
            nxt_id = id(nxt)
            if nxt_id not in node_id_map:
                node_id_map[nxt_id] = f"N{counter}"
                counter += 1
            tid = node_id_map[nxt_id]
            action_str = str(action or "default")
            ekey = (nid, tid, action_str)
            if ekey not in edge_seen:
                edge_seen.add(ekey)
                links.append({"source": nid, "target": tid, "action": action_str})
            queue.append(nxt)

    mermaid_lines = ["flowchart LR"]
    for n in nodes:
        mermaid_lines.append(f'    {n["id"]}["{n["label"]}"]')
    for e in links:
        mermaid_lines.append(f'    {e["source"]} -- "{e["action"]}" --> {e["target"]}')
    start_node_id = node_id_map.get(id(start), "")
    return {"nodes": nodes, "links": links, "mermaid": "\n".join(mermaid_lines), "start_node_id": start_node_id}


def _prompt_role_mapping() -> List[Dict[str, Any]]:
    role_map = [
        {"role": "Requirement Parser", "node": "ReqParseNode", "prompt": "req_parse.txt"},
        {"role": "Spec Lawyer", "node": "TriPersonaReviewNode", "prompt": "review_spec_lawyer.txt"},
        {"role": "Carrier Reviewer", "node": "TriPersonaReviewNode", "prompt": "review_carrier.txt"},
        {"role": "UX Advocate", "node": "TriPersonaReviewNode", "prompt": "review_ux_advocate.txt"},
        {"role": "Synthesis Arbiter", "node": "SynthesisNode", "prompt": "synthesis_arbiter.txt"},
        {"role": "Test Designer", "node": "TestDesignNode", "prompt": "test_designer.txt"},
        {"role": "Integrated Matrix Designer", "node": "TestDesignNode", "prompt": "integrated_matrix_designer.txt"},
        {"role": "Test Case Generator", "node": "TestCaseGeneratorNode", "prompt": "test_case_generator.txt"},
        {"role": "Harness Mapper", "node": "HarnessMapperNode", "prompt": "harness_mapper.txt"},
        {"role": "Script Writer", "node": "ScriptWriterNode", "prompt": "script_writer.txt"},
    ]
    out: List[Dict[str, Any]] = []
    for item in role_map:
        p = prompt_dir / item["prompt"]
        first_line = ""
        if p.exists():
            for ln in p.read_text(encoding="utf-8").splitlines():
                raw = ln.strip()
                if raw:
                    first_line = raw[:120]
                    break
        out.append(
            {
                **item,
                "exists": p.exists(),
                "first_line": first_line,
            }
        )
    return out


@app.get("/api/settings/flow-visualization")
def get_flow_visualization():
    flow = create_system_test_flow()
    graph = _flow_to_graph(flow)
    return {
        "flow_name": "SystemTestFlow",
        "mermaid": graph["mermaid"],
        "graph": {"nodes": graph["nodes"], "links": graph["links"], "start_node_id": graph.get("start_node_id", "")},
        "prompt_roles": _prompt_role_mapping(),
    }


@app.post("/api/settings/prompts")
def save_prompts(payload: PromptSettingsRequest):
    for name, content in payload.prompts.items():
        if name not in ALLOWED_PROMPT_FILES:
            raise HTTPException(status_code=400, detail=f"非法模板名: {name}")
        p = prompt_dir / name
        p.write_text(content, encoding="utf-8")
    return {"message": "saved", "count": len(payload.prompts)}


@app.post("/api/export/testcases.xlsx")
def export_testcases_excel(payload: ExportTestCasesRequest):
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"缺少openpyxl依赖: {exc}")

    requirement_input = payload.requirement_input or []
    requirement_spec = payload.requirement_spec or {}
    persona_reviews = payload.persona_reviews or {}
    requirement_review_history = payload.requirement_review_history or {}
    test_design_spec = payload.test_design_spec or {}
    test_case_spec = payload.test_case_spec or {}
    testcases = test_case_spec.get("testcases", [])

    def _list_to_text(v: Any) -> str:
        if isinstance(v, list):
            return "\n".join([str(x) for x in v])
        if v is None:
            return ""
        return str(v)

    def _safe_json(v: Any) -> str:
        try:
            return json.dumps(v, ensure_ascii=False)
        except Exception:
            return str(v)

    wb = Workbook()

    ws_input = wb.active
    ws_input.title = "需求输入"
    ws_input.append(["index", "requirement_text"])
    for i, txt in enumerate(requirement_input, start=1):
        ws_input.append([i, txt])

    ws_req = wb.create_sheet("测试需求稿")
    ws_req.append(["req_id", "title", "priority", "rat_scope", "persona_sources", "pass_fail", "kpi"])
    for r in requirement_spec.get("final_requirements", []) if isinstance(requirement_spec, dict) else []:
        if not isinstance(r, dict):
            continue
        ws_req.append(
            [
                r.get("req_id", ""),
                r.get("title", ""),
                r.get("priority", ""),
                _list_to_text(r.get("rat_scope", [])),
                _list_to_text(r.get("persona_sources", [])),
                _list_to_text((r.get("acceptance") or {}).get("pass_fail", [])),
                _list_to_text((r.get("acceptance") or {}).get("kpi", [])),
            ]
        )

    ws_review = wb.create_sheet("三人评审")
    ws_review.append(["persona", "req_id", "issues", "open_questions", "rewrite_suggestion", "scores_json"])
    for persona_key in ["spec", "carrier", "ux"]:
        revs = ((persona_reviews.get(persona_key) or {}).get("reviews", []) if isinstance(persona_reviews, dict) else [])
        for item in revs:
            if not isinstance(item, dict):
                continue
            ws_review.append(
                [
                    persona_key,
                    item.get("req_id", ""),
                    _list_to_text(item.get("issues", [])),
                    _list_to_text(item.get("open_questions", [])),
                    item.get("rewrite_suggestion", ""),
                    _safe_json(item.get("scores", {})),
                ]
            )

    ws_design_obj = wb.create_sheet("测试设计目标")
    ws_design_obj.append(["objective_id", "linked_reqs", "goal", "success_criteria", "evidence", "priority", "risk_notes"])
    for o in test_design_spec.get("objectives", []) if isinstance(test_design_spec, dict) else []:
        if not isinstance(o, dict):
            continue
        ws_design_obj.append(
            [
                o.get("objective_id", ""),
                _list_to_text(o.get("linked_reqs", [])),
                o.get("goal", ""),
                _list_to_text(o.get("success_criteria", [])),
                _list_to_text(o.get("evidence", [])),
                o.get("priority", ""),
                o.get("risk_notes", ""),
            ]
        )

    ws_design_matrix = wb.create_sheet("测试设计矩阵")
    ws_design_matrix.append(["row_id", "req_id", "objective_id", "scenario", "key_configuration", "pass_criteria", "recommended_test_environment", "availability_risk"])
    for row in test_design_spec.get("integrated_matrix", []) if isinstance(test_design_spec, dict) else []:
        if not isinstance(row, dict):
            continue
        env_text, risk_text = _format_test_environment_cell(row.get("test_environment"))
        ws_design_matrix.append(
            [
                row.get("row_id", ""),
                row.get("req_id", ""),
                row.get("objective_id", ""),
                row.get("scenario", ""),
                _safe_json(row.get("key_configuration", {})),
                _list_to_text(row.get("pass_criteria", [])),
                env_text,
                risk_text,
            ]
        )

    ws_cases = wb.create_sheet("测试用例")
    ws_cases.append(
        [
            "tc_id",
            "objective_id",
            "title",
            "tags",
            "preconditions",
            "expected",
            "pass_fail",
            "recommended_test_environment",
            "availability_risk",
            "steps_count",
            "must_capture",
        ]
    )
    for tc in testcases:
        if not isinstance(tc, dict):
            continue
        env_text, risk_text = _format_test_environment_cell(tc.get("test_environment"))
        ws_cases.append(
            [
                tc.get("tc_id", ""),
                tc.get("objective_id", ""),
                tc.get("title", ""),
                ", ".join(tc.get("tags", [])),
                _list_to_text(tc.get("preconditions", [])),
                _list_to_text(tc.get("expected", [])),
                _list_to_text(tc.get("pass_fail", [])),
                env_text,
                risk_text,
                len(tc.get("steps", [])),
                ", ".join((tc.get("observability") or {}).get("must_capture", [])),
            ]
        )

    ws_answers = wb.create_sheet("问题确认记录")
    ws_answers.append(["source_round", "req_id", "question", "answer"])
    for item in requirement_review_history.get("answered_questions", []) if isinstance(requirement_review_history, dict) else []:
        if not isinstance(item, dict):
            continue
        ws_answers.append(
            [
                item.get("source_round", ""),
                item.get("req_id", ""),
                item.get("question", ""),
                item.get("answer", ""),
            ]
        )

    ws_steps = wb.create_sheet("用例步骤")
    ws_steps.append(["tc_id", "step_index", "action", "params_json"])
    for tc in testcases:
        if not isinstance(tc, dict):
            continue
        tc_id = tc.get("tc_id", "")
        for idx, step in enumerate(tc.get("steps", []), start=1):
            if not isinstance(step, dict):
                continue
            ws_steps.append([tc_id, idx, step.get("action", ""), _safe_json(step.get("params", {}))])

    ws_raw = wb.create_sheet("RawJSON")
    ws_raw.append(["section", "json"])
    ws_raw.append(["requirement_spec", _safe_json(requirement_spec)])
    ws_raw.append(["persona_reviews", _safe_json(persona_reviews)])
    ws_raw.append(["requirement_review_history", _safe_json(requirement_review_history)])
    ws_raw.append(["test_design_spec", _safe_json(test_design_spec)])
    ws_raw.append(["test_case_spec", _safe_json(test_case_spec)])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    headers = {"Content-Disposition": "attachment; filename=testcases.xlsx"}
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def _run_cmd(args: List[str], timeout_s: int = 12) -> Dict[str, Any]:
    try:
        cp = subprocess.run(args, capture_output=True, text=True, timeout=timeout_s)
        return {
            "ok": cp.returncode == 0,
            "code": cp.returncode,
            "stdout": (cp.stdout or "")[-4000:],
            "stderr": (cp.stderr or "")[-2000:],
            "cmd": " ".join(args),
        }
    except Exception as exc:
        return {"ok": False, "code": -1, "stdout": "", "stderr": str(exc), "cmd": " ".join(args)}


def _pick_adb_device(preferred: str = "") -> str:
    res = _run_cmd(["adb", "devices"])
    if not res["ok"]:
        return preferred.strip()
    lines = (res.get("stdout") or "").splitlines()
    devices: List[str] = []
    for ln in lines:
        ln = ln.strip()
        if not ln or ln.lower().startswith("list of devices"):
            continue
        parts = ln.split()
        if len(parts) >= 2 and parts[1] == "device":
            devices.append(parts[0])
    if preferred and preferred in devices:
        return preferred
    return devices[0] if devices else preferred.strip()


def _list_adb_devices() -> Dict[str, Any]:
    res = _run_cmd(["adb", "devices"])
    devices: List[Dict[str, Any]] = []
    if not res.get("ok"):
        return {"ok": False, "devices": devices, "error": res.get("stderr", "adb执行失败")}

    for ln in (res.get("stdout") or "").splitlines():
        line = ln.strip()
        if not line or line.lower().startswith("list of devices"):
            continue
        parts = line.split()
        serial = parts[0] if parts else ""
        state = parts[1] if len(parts) > 1 else "unknown"
        if not serial:
            continue
        model = ""
        if state == "device":
            prop = _run_cmd(["adb", "-s", serial, "shell", "getprop", "ro.product.model"], timeout_s=6)
            model = (prop.get("stdout") or "").strip()
        devices.append({"id": serial, "state": state, "model": model})
    return {"ok": True, "devices": devices}


def _list_serial_ports() -> Dict[str, Any]:
    ports: List[Dict[str, Any]] = []
    try:
        import serial.tools.list_ports as list_ports  # type: ignore
    except Exception as exc:
        return {"ok": False, "ports": ports, "error": f"pyserial不可用: {exc}"}

    for p in list_ports.comports():
        ports.append(
            {
                "device": getattr(p, "device", ""),
                "name": getattr(p, "name", ""),
                "description": getattr(p, "description", ""),
                "hwid": getattr(p, "hwid", ""),
            }
        )
    return {"ok": True, "ports": ports}


@app.get("/api/devices/adb")
def devices_adb():
    return _list_adb_devices()


@app.get("/api/devices/serial")
def devices_serial():
    return _list_serial_ports()


def _adb_exec(device_id: str, shell_args: List[str], timeout_s: int = 15) -> Dict[str, Any]:
    dev = _pick_adb_device(device_id)
    if not dev:
        return {"ok": False, "cmd": "adb", "stderr": "未检测到可用Android设备", "stdout": ""}
    return _run_cmd(["adb", "-s", dev, "shell", *shell_args], timeout_s=timeout_s)


def _serial_exchange(port: str, baudrate: int, cmd: str, read_timeout: float = 1.2) -> Dict[str, Any]:
    if not port:
        return {"ok": False, "cmd": cmd, "stderr": "at_port为空", "stdout": ""}
    try:
        import serial  # type: ignore
    except Exception as exc:
        return {"ok": False, "cmd": cmd, "stderr": f"pyserial不可用: {exc}", "stdout": ""}

    try:
        with serial.Serial(port=port, baudrate=baudrate, timeout=read_timeout) as ser:
            ser.reset_input_buffer()
            ser.reset_output_buffer()
            ser.write((cmd.strip() + "\r").encode("utf-8"))
            time.sleep(0.2)
            data = ser.read(2048)
            out = data.decode(errors="ignore")
            ok = ("OK" in out.upper()) or ("+CREG" in out.upper()) or ("+COPS" in out.upper())
            return {"ok": ok, "cmd": cmd, "stdout": out[-4000:], "stderr": ""}
    except Exception as exc:
        return {"ok": False, "cmd": cmd, "stdout": "", "stderr": str(exc)}


@app.post("/api/automation/debug")
def automation_debug(payload: AutomationRequest):
    mode = (payload.mode or "").strip().lower()
    if mode not in {"android_adb", "at_serial"}:
        raise HTTPException(status_code=400, detail="mode必须是 android_adb 或 at_serial")

    checks: List[Dict[str, Any]] = []
    if mode == "android_adb":
        checks.append(_run_cmd(["adb", "version"]))
        dev = _pick_adb_device(payload.device_id)
        checks.append(_run_cmd(["adb", "devices"]))
        if dev:
            checks.append(_run_cmd(["adb", "-s", dev, "get-state"]))
            checks.append(_run_cmd(["adb", "-s", dev, "shell", "getprop", "ro.product.model"]))
        else:
            checks.append({"ok": False, "cmd": "adb get-state", "stdout": "", "stderr": "未找到在线设备"})
    else:
        checks.append(_serial_exchange(payload.at_port, payload.baudrate, "AT"))
        checks.append(_serial_exchange(payload.at_port, payload.baudrate, "AT+CREG?"))

    ok = all(bool(x.get("ok")) for x in checks)
    return {"ok": ok, "mode": mode, "checks": checks}


def _run_android_testcase(tc: Dict[str, Any], device_id: str, max_steps: int) -> Dict[str, Any]:
    steps = tc.get("steps", []) if isinstance(tc.get("steps", []), list) else []
    steps = steps[: max(1, max_steps)]
    step_results: List[Dict[str, Any]] = []
    for i, st in enumerate(steps, start=1):
        action = str(st.get("action", "")).upper()
        params = st.get("params", {}) if isinstance(st.get("params", {}), dict) else {}
        if action == "SET_DEVICE_STATE":
            airplane = params.get("airplane")
            if airplane is True:
                res = _adb_exec(device_id, ["cmd", "connectivity", "airplane-mode", "enable"])
            elif airplane is False:
                res = _adb_exec(device_id, ["cmd", "connectivity", "airplane-mode", "disable"])
            else:
                res = _adb_exec(device_id, ["getprop", "ro.build.version.release"])
        elif action == "TRIGGER_REGISTRATION":
            res = _adb_exec(device_id, ["svc", "data", "enable"])
        elif action == "WAIT":
            time.sleep(min(int(params.get("timeout_s", 1)), 5))
            res = {"ok": True, "cmd": "WAIT", "stdout": "sleep done", "stderr": ""}
        elif action == "COLLECT_LOG":
            dev = _pick_adb_device(device_id)
            res = _run_cmd(["adb", "-s", dev, "logcat", "-d", "-t", "120"]) if dev else {"ok": False, "cmd": "adb logcat", "stdout": "", "stderr": "无设备"}
        else:
            res = _adb_exec(device_id, ["dumpsys", "telephony.registry"], timeout_s=18)
        step_results.append({"step_index": i, "action": action, **res})
    passed = all(x.get("ok") for x in step_results)
    return {"tc_id": tc.get("tc_id", ""), "passed": passed, "steps": step_results}


def _run_at_testcase(tc: Dict[str, Any], port: str, baudrate: int, max_steps: int) -> Dict[str, Any]:
    steps = tc.get("steps", []) if isinstance(tc.get("steps", []), list) else []
    steps = steps[: max(1, max_steps)]
    step_results: List[Dict[str, Any]] = []
    for i, st in enumerate(steps, start=1):
        action = str(st.get("action", "")).upper()
        if action in {"TRIGGER_REGISTRATION", "CHECK"}:
            cmd = "AT+CREG?"
        elif action == "SET_DEVICE_STATE":
            cmd = "AT+CFUN=1"
        else:
            cmd = "AT"
        res = _serial_exchange(port, baudrate, cmd)
        step_results.append({"step_index": i, "action": action, **res})
    passed = all(x.get("ok") for x in step_results)
    return {"tc_id": tc.get("tc_id", ""), "passed": passed, "steps": step_results}


@app.post("/api/automation/run")
def automation_run(payload: AutomationRequest):
    mode = (payload.mode or "").strip().lower()
    if mode not in {"android_adb", "at_serial"}:
        raise HTTPException(status_code=400, detail="mode必须是 android_adb 或 at_serial")
    spec = payload.test_case_spec or {}
    testcases = spec.get("testcases", [])
    if not isinstance(testcases, list) or not testcases:
        raise HTTPException(status_code=400, detail="缺少test_case_spec.testcases")

    max_cases = max(1, min(payload.max_cases, 30))
    max_steps = max(1, min(payload.max_steps, 40))
    selected = testcases[:max_cases]
    results: List[Dict[str, Any]] = []
    for tc in selected:
        if not isinstance(tc, dict):
            continue
        if mode == "android_adb":
            results.append(_run_android_testcase(tc, payload.device_id, max_steps))
        else:
            results.append(_run_at_testcase(tc, payload.at_port, payload.baudrate, max_steps))

    total = len(results)
    passed = sum(1 for r in results if r.get("passed"))
    return {
        "mode": mode,
        "summary": {"total": total, "passed": passed, "failed": total - passed},
        "results": results,
    }


def _load_json_file(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _render_at_command(template: str, params: Dict[str, Any]) -> str:
    cmd = template or ""
    p = params or {}
    for k, v in p.items():
        cmd = cmd.replace("{" + str(k) + "}", str(v))
    # keep unresolved placeholders empty for MVP
    while "{" in cmd and "}" in cmd:
        s = cmd.find("{")
        e = cmd.find("}", s + 1)
        if s < 0 or e < 0:
            break
        cmd = cmd[:s] + "" + cmd[e + 1 :]
    return cmd


def _exec_at_command(mode: str, cmd: str, payload: AutomationRequest) -> Dict[str, Any]:
    if mode == "at_serial":
        return _serial_exchange(payload.at_port, payload.baudrate, cmd)
    if mode == "android_adb":
        # Experimental fallback: log the AT intent and collect telephony snapshot via adb.
        snapshot = _adb_exec(payload.device_id, ["dumpsys", "telephony.registry"], timeout_s=15)
        return {
            "ok": bool(snapshot.get("ok")),
            "cmd": cmd,
            "stdout": snapshot.get("stdout", ""),
            "stderr": snapshot.get("stderr", ""),
            "note": "android_adb 模式下 AT 命令为实验映射（非直连modem口）",
        }
    return {"ok": False, "cmd": cmd, "stdout": "", "stderr": f"unsupported mode: {mode}"}


def _find_transition_plan(efsm: Dict[str, Any], max_steps: int) -> List[Dict[str, Any]]:
    transitions = efsm.get("transitions", [])
    if not isinstance(transitions, list):
        return []
    # MVP heuristic: prioritize unique coverage points, then original order.
    # Global edges (from='*' / to='*') are recovery edges and should not dominate
    # normal coverage runs.
    normal: List[Dict[str, Any]] = []
    global_edges: List[Dict[str, Any]] = []
    for t in transitions:
        if not isinstance(t, dict):
            continue
        if str(t.get("from", "")) == "*" or str(t.get("to", "")) == "*":
            global_edges.append(t)
        else:
            normal.append(t)
    ordered = normal + global_edges

    seen_cov = set()
    plan: List[Dict[str, Any]] = []
    for t in ordered:
        if not isinstance(t, dict):
            continue
        cov = (t.get("coverage", {}) or {}).get("points", []) if isinstance(t.get("coverage", {}), dict) else []
        key = tuple(cov) if isinstance(cov, list) else tuple()
        if key and key in seen_cov:
            continue
        if key:
            seen_cov.add(key)
        plan.append(t)
        if len(plan) >= max_steps:
            break
    return plan


def _merge_atspec(base: Dict[str, Any], ext: Dict[str, Any]) -> Dict[str, Any]:
    out = json.loads(json.dumps(base))
    out_caps = out.setdefault("capabilities", [])
    out_cmds = out.setdefault("commands", [])
    cap_ids = {c.get("id") for c in out_caps if isinstance(c, dict)}
    cmd_ids = {c.get("id") for c in out_cmds if isinstance(c, dict)}
    for c in (ext.get("capabilities", []) if isinstance(ext.get("capabilities", []), list) else []):
        if isinstance(c, dict) and c.get("id") and c.get("id") not in cap_ids:
            out_caps.append(c)
            cap_ids.add(c.get("id"))
    for c in (ext.get("commands", []) if isinstance(ext.get("commands", []), list) else []):
        if isinstance(c, dict) and c.get("id") and c.get("id") not in cmd_ids:
            out_cmds.append(c)
            cmd_ids.add(c.get("id"))
    return out


def _extract_token_from_template(s: str) -> str:
    text = (s or "").strip().upper()
    m = re.match(r"^AT(?:\+)?([A-Z0-9]+)", text)
    return m.group(1) if m else ""


def _command_template(cmd: Dict[str, Any]) -> str:
    if not isinstance(cmd, dict):
        return ""
    if isinstance(cmd.get("at"), str):
        return cmd.get("at", "")
    forms = cmd.get("forms", {})
    if isinstance(forms, dict):
        for k in ["set", "read", "test", "exec", "exec_interactive", "action"]:
            if isinstance(forms.get(k), str):
                return forms[k]
        for v in forms.values():
            if isinstance(v, str):
                return v
    return ""


def _json_pointer_unescape(token: str) -> str:
    return token.replace("~1", "/").replace("~0", "~")


def _resolve_pointer_parent(doc: Any, pointer: str):
    if pointer == "":
        return None, None
    if not pointer.startswith("/"):
        raise ValueError(f"invalid JSON pointer: {pointer}")
    parts = [_json_pointer_unescape(p) for p in pointer.split("/")[1:]]
    if not parts:
        return None, None
    parent = doc
    for p in parts[:-1]:
        if isinstance(parent, dict):
            if p not in parent:
                parent[p] = {}
            parent = parent[p]
        elif isinstance(parent, list):
            if p == "-":
                raise ValueError("invalid pointer '-' in middle")
            idx = int(p)
            if idx < 0 or idx >= len(parent):
                raise ValueError(f"list index out of range: {idx}")
            parent = parent[idx]
        else:
            raise ValueError(f"cannot traverse pointer segment: {p}")
    return parent, parts[-1]


def _set_by_pointer(doc: Any, pointer: str, value: Any, create_only: bool = False):
    if pointer == "":
        return value
    parent, token = _resolve_pointer_parent(doc, pointer)
    if isinstance(parent, dict):
        if create_only and token in parent:
            raise ValueError(f"path already exists: {pointer}")
        parent[token] = value
        return doc
    if isinstance(parent, list):
        if token == "-":
            parent.append(value)
            return doc
        idx = int(token)
        if create_only and idx < len(parent):
            raise ValueError(f"path already exists: {pointer}")
        if idx < 0 or idx > len(parent):
            raise ValueError(f"list index out of range: {idx}")
        if idx == len(parent):
            parent.append(value)
        else:
            parent[idx] = value
        return doc
    raise ValueError(f"invalid parent for pointer: {pointer}")


def _remove_by_pointer(doc: Any, pointer: str):
    parent, token = _resolve_pointer_parent(doc, pointer)
    if isinstance(parent, dict):
        if token in parent:
            del parent[token]
            return
        raise ValueError(f"path not found: {pointer}")
    if isinstance(parent, list):
        idx = int(token)
        if idx < 0 or idx >= len(parent):
            raise ValueError(f"list index out of range: {idx}")
        del parent[idx]
        return
    raise ValueError(f"invalid parent for pointer: {pointer}")


def _get_by_pointer(doc: Any, pointer: str):
    if pointer == "":
        return doc
    if not pointer.startswith("/"):
        raise ValueError(f"invalid JSON pointer: {pointer}")
    cur = doc
    for part in pointer.split("/")[1:]:
        token = _json_pointer_unescape(part)
        if isinstance(cur, dict):
            if token not in cur:
                raise ValueError(f"path not found: {pointer}")
            cur = cur[token]
        elif isinstance(cur, list):
            if token == "-":
                raise ValueError("invalid pointer '-' in get")
            idx = int(token)
            if idx < 0 or idx >= len(cur):
                raise ValueError(f"list index out of range: {idx}")
            cur = cur[idx]
        else:
            raise ValueError(f"invalid path segment: {token}")
    return cur


def _apply_json_patch(doc: Any, patch_ops: Any) -> Any:
    if not isinstance(patch_ops, list):
        raise ValueError("JSON Patch must be a list")
    out = json.loads(json.dumps(doc))
    for op in patch_ops:
        if not isinstance(op, dict):
            raise ValueError("patch op must be object")
        kind = str(op.get("op", "")).strip()
        path = str(op.get("path", "")).strip()
        if kind not in {"add", "remove", "replace", "test"}:
            raise ValueError(f"unsupported patch op: {kind}")
        if kind == "add":
            out = _set_by_pointer(out, path, op.get("value"), create_only=False)
        elif kind == "replace":
            _get_by_pointer(out, path)
            out = _set_by_pointer(out, path, op.get("value"), create_only=False)
        elif kind == "remove":
            _remove_by_pointer(out, path)
        elif kind == "test":
            cur = _get_by_pointer(out, path)
            if cur != op.get("value"):
                raise ValueError(f"test op failed at {path}")
    return out


def _validate_manifest(doc: Dict[str, Any]):
    if not isinstance(doc, dict):
        raise ValueError("manifest must be object")
    if not isinstance(doc.get("baseline"), str) or not doc.get("baseline"):
        raise ValueError("manifest.baseline must be non-empty string")
    if "extensions" in doc and not isinstance(doc.get("extensions"), list):
        raise ValueError("manifest.extensions must be array")
    policy = doc.get("policy", {})
    if policy and not isinstance(policy, dict):
        raise ValueError("manifest.policy must be object")
    for k in ["must_have_capabilities", "allowed_missing_capabilities"]:
        if k in policy and not isinstance(policy.get(k), list):
            raise ValueError(f"manifest.policy.{k} must be array")
    scope = doc.get("test_scope", {})
    if scope and not isinstance(scope, dict):
        raise ValueError("manifest.test_scope must be object")
    for k in ["enable_capabilities", "disable_capabilities", "enable_commands", "disable_commands"]:
        if k in scope and not isinstance(scope.get(k), list):
            raise ValueError(f"manifest.test_scope.{k} must be array")
    env = doc.get("env", {})
    if env and not isinstance(env, dict):
        raise ValueError("manifest.env must be object")


def _validate_extension(doc: Dict[str, Any]):
    if not isinstance(doc, dict):
        raise ValueError("extension must be object")
    meta = doc.get("meta", {})
    if not isinstance(meta, dict):
        raise ValueError("extension.meta must be object")
    if not isinstance(meta.get("id"), str) or not meta.get("id"):
        raise ValueError("extension.meta.id must be non-empty string")
    if not isinstance(meta.get("version"), str) or not meta.get("version"):
        raise ValueError("extension.meta.version must be non-empty string")
    caps = doc.get("capabilities", [])
    if not isinstance(caps, list):
        raise ValueError("extension.capabilities must be array")
    for i, c in enumerate(caps):
        if not isinstance(c, dict) or not c.get("id"):
            raise ValueError(f"extension.capabilities[{i}] invalid")
    cmds = doc.get("commands", [])
    if not isinstance(cmds, list):
        raise ValueError("extension.commands must be array")
    for i, c in enumerate(cmds):
        if not isinstance(c, dict) or not c.get("id"):
            raise ValueError(f"extension.commands[{i}] missing id")
        forms = c.get("forms", {})
        if forms and not isinstance(forms, dict):
            raise ValueError(f"extension.commands[{i}].forms must be object")
        responses = c.get("responses", {})
        if responses and not isinstance(responses, dict):
            raise ValueError(f"extension.commands[{i}].responses must be object")


def _normalize_manifest_by_request(manifest: Dict[str, Any], request_text: str) -> Dict[str, Any]:
    out = json.loads(json.dumps(manifest or {}))
    scope = out.setdefault("test_scope", {})
    if not isinstance(scope, dict):
        scope = {}
        out["test_scope"] = scope
    disable_caps = scope.get("disable_capabilities", [])
    if not isinstance(disable_caps, list):
        disable_caps = []
    # dedupe while preserving order
    seen = set()
    normalized = []
    for x in disable_caps:
        s = str(x).strip()
        if not s or s in seen:
            continue
        seen.add(s)
        normalized.append(s)

    req = (request_text or "").lower()
    explicit_cs_reg = any(k in req for k in ["net.registration.cs", "cs注册", "creg", "+creg", "cs registration"])
    if "net.registration.cs" in normalized and not explicit_cs_reg:
        normalized = [x for x in normalized if x != "net.registration.cs"]
    if any(k in req for k in ["cs语音", "语音", "通话", "voice", "call"]) and "cs.call.basic" not in normalized:
        normalized.append("cs.call.basic")
    if any(k in req for k in ["sms", "短信"]) and all(x not in normalized for x in ["sms.basic", "sms.core"]):
        normalized.append("sms.basic")
    scope["disable_capabilities"] = normalized
    return out


def _extract_json_block(text: str) -> Any:
    try:
        return json.loads(text)
    except Exception:
        pass
    m = re.search(r"```(?:json)?\s*(.*?)```", text, flags=re.S)
    if m:
        return json.loads(m.group(1).strip())
    raise ValueError("no valid JSON found in LLM output")


def _build_config_compile_prompt(
    request_text: str, manifest: Dict[str, Any], extension: Dict[str, Any], error_hint: str = ""
) -> str:
    payload = {
        "user_request": request_text,
        "current_manifest": manifest,
        "current_extension": extension,
        "output_contract": {
            "change_spec": {
                "manifest_changes": {
                    "disable_capabilities": ["string"],
                    "disable_commands": ["string"],
                    "add_extensions": ["string"],
                    "env_set": {"key": "value"},
                },
                "extension_requests": [{"vendor": "string", "capability": "string", "command_id": "string"}],
            },
            "manifest_patch": [{"op": "add|remove|replace|test", "path": "/json/pointer", "value": "any"}],
            "extension_mode": "replace|patch",
            "extension_patch": [{"op": "add|remove|replace|test", "path": "/json/pointer", "value": "any"}],
            "extension_file": {"meta": {"id": "string", "version": "string"}, "capabilities": [], "commands": []},
        },
        "rules": [
            "Output ONLY JSON object, no markdown.",
            "Do not remove baseline or existing extensions unless user explicitly asks.",
            "Use minimal patch, avoid unrelated changes.",
            "If extension changes are needed, prefer extension_mode='replace' with full valid extension_file.",
        ],
        "error_hint_from_previous_attempt": error_hint,
    }
    return json.dumps(payload, ensure_ascii=False)


def _generate_config_updates_with_retry(
    request_text: str, use_llm: bool = True, max_attempts: int = 3
) -> Dict[str, Any]:
    manifest = _load_json_file(at_manifest_path)
    extension = _load_json_file(at_extension_path)
    last_error = ""
    output: Dict[str, Any] = {}

    for attempt in range(1, max_attempts + 1):
        if use_llm:
            client = LLMClient()
            if not client.enabled:
                raise ValueError("LLM not configured")
            raw = client.generate(
                system_prompt=(
                    "You are ConfigPatchBot. Return only strict JSON object. "
                    "Generate ChangeSpec + JSON Patch + Extension update draft."
                ),
                user_prompt=_build_config_compile_prompt(request_text, manifest, extension, error_hint=last_error),
                temperature=0.1,
            )
            output = _extract_json_block(raw)
        else:
            output = {
                "change_spec": {
                    "manifest_changes": {
                        "disable_capabilities": [],
                        "disable_commands": [],
                        "add_extensions": [],
                        "env_set": {},
                    },
                    "extension_requests": [],
                },
                "manifest_patch": [],
                "extension_mode": "replace",
                "extension_file": extension,
            }

        try:
            m_patch = output.get("manifest_patch", [])
            if not isinstance(m_patch, list):
                raise ValueError("manifest_patch must be array")
            manifest_new = _apply_json_patch(manifest, m_patch) if m_patch else manifest
            manifest_new = _normalize_manifest_by_request(manifest_new, request_text)

            ext_mode = str(output.get("extension_mode", "replace"))
            if ext_mode == "patch":
                e_patch = output.get("extension_patch", [])
                if not isinstance(e_patch, list):
                    raise ValueError("extension_patch must be array")
                extension_new = _apply_json_patch(extension, e_patch) if e_patch else extension
            else:
                extension_new = output.get("extension_file", extension)

            _validate_manifest(manifest_new)
            _validate_extension(extension_new)
            return {
                "change_spec": output.get("change_spec", {}),
                "manifest_patch": m_patch,
                "extension_mode": ext_mode,
                "extension_patch": output.get("extension_patch", []),
                "manifest_new": manifest_new,
                "extension_new": extension_new,
                "attempts": attempt,
                "validation_error": "",
            }
        except Exception as exc:
            last_error = str(exc)
            if attempt >= max_attempts:
                raise ValueError(f"config compile failed after retries: {last_error}")

    raise ValueError("config compile failed")


def _compile_at_assets(use_llm: bool = True) -> Dict[str, Any]:
    baseline = _load_json_file(at_spec_path)
    extension = _load_json_file(at_extension_path)
    profile = _load_json_file(at_profile_path)
    manifest = _load_json_file(at_manifest_path)
    efsm_template = _load_json_file(at_efsm_path)

    effective_atspec = _merge_atspec(baseline, extension)
    test_scope = manifest.get("test_scope", {}) if isinstance(manifest.get("test_scope", {}), dict) else {}
    disable_caps = set(
        test_scope.get("disable_capabilities", []) if isinstance(test_scope.get("disable_capabilities", []), list) else []
    )
    enable_caps = set(test_scope.get("enable_capabilities", []) if isinstance(test_scope.get("enable_capabilities", []), list) else ["*"])
    disable_cmds = set(test_scope.get("disable_commands", []) if isinstance(test_scope.get("disable_commands", []), list) else [])
    enable_cmds = set(test_scope.get("enable_commands", []) if isinstance(test_scope.get("enable_commands", []), list) else [])

    raw_caps = effective_atspec.get("capabilities", []) if isinstance(effective_atspec.get("capabilities", []), list) else []

    def _cap_family(cap_id: str) -> str:
        c = (cap_id or "").lower()
        if c in {"sms.basic", "sms.core"} or c.startswith("sms."):
            return "sms"
        if c in {"cs.call.basic", "cs.call"} or c.startswith("cs.call.") or "voice" in c:
            return "cs.call"
        return c

    disabled_families = {_cap_family(str(x)) for x in disable_caps}

    def _cap_disabled(cap_id: str) -> bool:
        c = str(cap_id or "")
        if c in disable_caps:
            return True
        return _cap_family(c) in disabled_families
    if "*" in enable_caps:
        caps_kept = [c for c in raw_caps if isinstance(c, dict) and c.get("id") and not _cap_disabled(str(c.get("id")))]
    else:
        caps_kept = [
            c
            for c in raw_caps
            if isinstance(c, dict) and c.get("id") in enable_caps and not _cap_disabled(str(c.get("id")))
        ]
    cap_ids_kept = {c.get("id") for c in caps_kept if isinstance(c, dict)}
    effective_atspec["capabilities"] = caps_kept

    raw_cmds = effective_atspec.get("commands", []) if isinstance(effective_atspec.get("commands", []), list) else []
    cmd_objs: List[Dict[str, Any]] = []
    for c in raw_cmds:
        if not isinstance(c, dict) or not c.get("id"):
            continue
        cid = str(c.get("id"))
        if enable_cmds and cid not in enable_cmds:
            continue
        if cid in disable_cmds:
            continue
        cap = c.get("capability")
        if cap and (cap not in cap_ids_kept or _cap_disabled(str(cap))):
            continue
        cmd_objs.append(c)
    effective_atspec["commands"] = cmd_objs

    token_to_cap_all: Dict[str, str] = {}
    for c in raw_cmds:
        if not isinstance(c, dict):
            continue
        token = _extract_token_from_template(_command_template(c))
        if token and c.get("capability"):
            token_to_cap_all[token] = str(c.get("capability"))

    token_to_cap: Dict[str, str] = {}
    cmd_ids_enabled = {str(c.get("id")) for c in cmd_objs}
    cmd_tokens_enabled = set()
    for c in cmd_objs:
        cid = str(c.get("id"))
        token = _extract_token_from_template(_command_template(c))
        if token:
            cmd_tokens_enabled.add(token)
            if c.get("capability"):
                token_to_cap[token] = str(c.get("capability"))
        token_to_cap[cid.upper()] = str(c.get("capability") or "")

    effective_profile = json.loads(json.dumps(profile))
    bindings = effective_profile.get("bindings", []) if isinstance(effective_profile.get("bindings", []), list) else []
    new_bindings = []
    for b in bindings:
        if not isinstance(b, dict):
            continue
        cap = b.get("capability")
        if cap and (cap not in cap_ids_kept or _cap_disabled(str(cap))):
            continue
        impl = b.get("impl", []) if isinstance(b.get("impl", []), list) else []
        impl2 = [x for x in impl if isinstance(x, dict) and (not x.get("cmd_id") or str(x.get("cmd_id")) in cmd_ids_enabled)]
        b2 = dict(b)
        b2["impl"] = impl2
        new_bindings.append(b2)
    effective_profile["bindings"] = new_bindings
    effective_profile["capability_support"] = {
        c: any((isinstance(b, dict) and b.get("capability") == c and b.get("impl")) for b in new_bindings)
        for c in sorted(cap_ids_kept)
    }

    active_efsm = json.loads(json.dumps(efsm_template))
    transitions = active_efsm.get("transitions", []) if isinstance(active_efsm.get("transitions", []), list) else []
    pruned: List[Dict[str, Any]] = []
    kept: List[Dict[str, Any]] = []
    for t in transitions:
        if not isinstance(t, dict):
            continue
        reason = ""
        action = t.get("action", {}) if isinstance(t.get("action", {}), dict) else {}
        cap = action.get("capability")
        if cap and cap not in cap_ids_kept:
            reason = f"capability_disabled:{cap}"
        if not reason and action.get("cmd_id"):
            cid = str(action.get("cmd_id"))
            if cid not in cmd_ids_enabled:
                reason = f"command_disabled:{cid}"
        if not reason and isinstance(action.get("cmd_sequence"), list):
            for cid in action.get("cmd_sequence", []):
                if str(cid) not in cmd_ids_enabled:
                    reason = f"command_disabled:{cid}"
                    break
        if not reason:
            steps = action.get("steps", []) if isinstance(action.get("steps", []), list) else []
            for s in steps:
                if not isinstance(s, dict):
                    continue
                send = s.get("send", "")
                if not send and isinstance(s.get("cmd", {}), dict):
                    send = s.get("cmd", {}).get("send", "")
                token = _extract_token_from_template(str(send))
                if token:
                    cap_for_token = token_to_cap_all.get(token, "")
                    if cap_for_token and _cap_disabled(cap_for_token):
                        reason = f"step_token_capability_disabled:{token}:{cap_for_token}"
                        break
                    if token in token_to_cap and token not in cmd_tokens_enabled:
                        reason = f"step_token_disabled:{token}"
                        break
        if not reason and disabled_families:
            tid = str(t.get("id", "")).upper()
            frm = str(t.get("from", "")).upper()
            to = str(t.get("to", "")).upper()
            family_hit = ""
            if any("SMS" in x for x in [tid, frm, to]):
                family_hit = "sms"
            elif any(k in tid for k in ["VOICE", "CALL", "DIAL", "CHUP", "ATA", "ATH", "RING", "CLCC"]):
                family_hit = "cs.call"
            if family_hit and family_hit in disabled_families:
                reason = f"transition_disabled_family:{family_hit}"
        if reason:
            pruned.append({"transition_id": t.get("id"), "reason": reason})
        else:
            kept.append(t)
    active_efsm["transitions"] = kept

    # Keep only states still used by remaining transitions (plus initial states).
    all_states = active_efsm.get("states", []) if isinstance(active_efsm.get("states", []), list) else []
    used_state_ids = set()
    for t in kept:
        if not isinstance(t, dict):
            continue
        frm = str(t.get("from", ""))
        to = str(t.get("to", ""))
        if frm and frm != "*":
            used_state_ids.add(frm)
        if to and to != "*":
            used_state_ids.add(to)
    for s in all_states:
        if isinstance(s, dict) and str(s.get("type", "")).lower() == "initial" and s.get("id"):
            used_state_ids.add(str(s.get("id")))
    if used_state_ids:
        active_efsm["states"] = [s for s in all_states if isinstance(s, dict) and str(s.get("id", "")) in used_state_ids]

    policy = manifest.get("policy", {}) if isinstance(manifest.get("policy", {}), dict) else {}
    must_caps = policy.get("must_have_capabilities", []) if isinstance(policy.get("must_have_capabilities", []), list) else []
    allowed_missing = set(policy.get("allowed_missing_capabilities", []) if isinstance(policy.get("allowed_missing_capabilities", []), list) else [])
    unsupported = [c for c in must_caps if not effective_profile["capability_support"].get(c, False)]
    compliance_blockers = [c for c in unsupported if c not in allowed_missing]

    warnings = []
    if pruned:
        warnings.append(f"pruned_transitions={len(pruned)}")
    if compliance_blockers:
        warnings.append(f"must_have_missing={','.join(compliance_blockers)}")

    llm_advice: Dict[str, Any] = {"enabled": bool(use_llm), "summary": ""}
    if use_llm:
        client = LLMClient()
        if client.enabled:
            try:
                prompt = (
                    "你是AT MBT编译审查助手。请基于以下编译结果给出简明建议（JSON，字段: risks[], repairs[], profile_patch_hints[]）。\\n"
                    + json.dumps(
                        {
                            "warnings": warnings,
                            "pruned": pruned[:30],
                            "unsupported_must_caps": compliance_blockers,
                            "manifest_scope": test_scope,
                        },
                        ensure_ascii=False,
                    )
                )
                text = client.generate(
                    system_prompt="Return only compact JSON.",
                    user_prompt=prompt,
                    temperature=0.2,
                )
                try:
                    llm_advice = json.loads(text)
                    llm_advice["enabled"] = True
                except Exception:
                    llm_advice = {"enabled": True, "summary": text[:1500]}
            except Exception as exc:
                llm_advice = {"enabled": True, "summary": f"llm_error: {exc}"}
        else:
            llm_advice = {"enabled": False, "summary": "LLM not configured"}

    report = {
        "compiler": "ATSpec/EFSM Compiler v0.1",
        "warnings": warnings,
        "pruned_transitions": pruned,
        "unsupported_must_have_capabilities": compliance_blockers,
        "allowed_missing_capabilities": [c for c in unsupported if c in allowed_missing],
        "stats": {
            "capabilities": len(cap_ids_kept),
            "commands": len(cmd_ids_enabled),
            "transitions_before": len(transitions),
            "transitions_after": len(kept),
        },
        "llm_advice": llm_advice,
    }

    at_effective_atspec_path.write_text(json.dumps(effective_atspec, ensure_ascii=False, indent=2), encoding="utf-8")
    at_effective_profile_path.write_text(json.dumps(effective_profile, ensure_ascii=False, indent=2), encoding="utf-8")
    at_active_efsm_path.write_text(json.dumps(active_efsm, ensure_ascii=False, indent=2), encoding="utf-8")
    at_compile_report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    return {
        "effective_atspec": effective_atspec,
        "effective_profile": effective_profile,
        "active_efsm": active_efsm,
        "report": report,
    }


@app.get("/api/at-agent/spec")
def at_agent_get_spec():
    return _load_json_file(at_spec_path)


@app.post("/api/at-agent/spec")
def at_agent_save_spec(payload: AtAssetSaveRequest):
    # ATSpec baseline is treated as normative/locked in UI. Keep endpoint for controlled updates.
    if payload.locked_baseline:
        existing = _load_json_file(at_spec_path)
        if existing:
            # merge-only save: preserve baseline core, allow additive updates
            merged = _merge_atspec(existing, payload.data)
            # keep canonical baseline id if already present
            if existing.get("meta"):
                merged.setdefault("meta", {})
                merged["meta"]["id"] = existing.get("meta", {}).get("id", merged.get("meta", {}).get("id"))
            at_spec_path.write_text(json.dumps(merged, ensure_ascii=False, indent=2), encoding="utf-8")
            return {"saved": True, "path": str(at_spec_path), "mode": "merge_locked"}
    at_spec_path.write_text(json.dumps(payload.data, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"saved": True, "path": str(at_spec_path)}


@app.get("/api/at-agent/profile")
def at_agent_get_profile():
    return _load_json_file(at_profile_path)


@app.post("/api/at-agent/profile")
def at_agent_save_profile(payload: AtAssetSaveRequest):
    at_profile_path.write_text(json.dumps(payload.data, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"saved": True, "path": str(at_profile_path)}


@app.get("/api/at-agent/manifest")
def at_agent_get_manifest():
    return _load_json_file(at_manifest_path)


@app.post("/api/at-agent/manifest")
def at_agent_save_manifest(payload: AtAssetSaveRequest):
    at_manifest_path.write_text(json.dumps(payload.data, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"saved": True, "path": str(at_manifest_path)}


@app.get("/api/at-agent/extension")
def at_agent_get_extension():
    return _load_json_file(at_extension_path)


@app.post("/api/at-agent/extension")
def at_agent_save_extension(payload: AtAssetSaveRequest):
    at_extension_path.write_text(json.dumps(payload.data, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"saved": True, "path": str(at_extension_path)}


@app.get("/api/at-agent/efsm")
def at_agent_get_efsm():
    return _load_json_file(at_efsm_path)


@app.post("/api/at-agent/efsm")
def at_agent_save_efsm(payload: AtAssetSaveRequest):
    at_efsm_path.write_text(json.dumps(payload.data, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"saved": True, "path": str(at_efsm_path)}


@app.post("/api/at-agent/compile")
def at_agent_compile(payload: AtCompileRequest):
    try:
        return _compile_at_assets(use_llm=payload.use_llm)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"AT编译失败: {exc}")


@app.post("/api/at-agent/config-compile")
def at_agent_config_compile(payload: AtConfigCompileRequest):
    try:
        result = _generate_config_updates_with_retry(
            request_text=payload.request_text.strip(),
            use_llm=payload.use_llm,
            max_attempts=3,
        )
        manifest_new = result["manifest_new"]
        extension_new = result["extension_new"]

        if payload.apply_changes:
            at_manifest_path.write_text(json.dumps(manifest_new, ensure_ascii=False, indent=2), encoding="utf-8")
            at_extension_path.write_text(json.dumps(extension_new, ensure_ascii=False, indent=2), encoding="utf-8")

        compiled = {}
        if payload.apply_changes and payload.compile_after_apply:
            compiled = _compile_at_assets(use_llm=payload.use_llm)

        return {
            "applied": bool(payload.apply_changes),
            "compiled": bool(payload.apply_changes and payload.compile_after_apply),
            "change_spec": result.get("change_spec", {}),
            "manifest_patch": result.get("manifest_patch", []),
            "extension_mode": result.get("extension_mode", "replace"),
            "extension_patch": result.get("extension_patch", []),
            "manifest": manifest_new,
            "extension": extension_new,
            "retries_used": result.get("attempts", 1),
            "compile_result": compiled,
        }
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"AT配置编译失败: {exc}")


@app.post("/api/at-agent/reset-baseline")
def at_agent_reset_baseline():
    try:
        # Keep normative baseline ATSpec/Profile/EFSM files as-is,
        # reset user-mutated config layers to defaults.
        at_manifest_path.write_text(json.dumps(ATMANIFEST_DEFAULT, ensure_ascii=False, indent=2), encoding="utf-8")
        at_extension_path.write_text(json.dumps(ATEXTENSION_DEFAULT, ensure_ascii=False, indent=2), encoding="utf-8")
        compiled = _compile_at_assets(use_llm=False)
        return {
            "reset": True,
            "manifest": _load_json_file(at_manifest_path),
            "extension": _load_json_file(at_extension_path),
            "compile_result": compiled,
        }
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"AT基线重置失败: {exc}")


@app.get("/api/at-agent/build")
def at_agent_build_get():
    return {
        "effective_atspec": _load_json_file(at_effective_atspec_path),
        "effective_profile": _load_json_file(at_effective_profile_path),
        "active_efsm": _load_json_file(at_active_efsm_path),
        "report": _load_json_file(at_compile_report_path),
    }


@app.post("/api/at-agent/mbt/run")
def at_agent_mbt_run(payload: AutomationRequest):
    mode = (payload.mode or "").strip().lower()
    if mode not in {"at_serial", "android_adb"}:
        raise HTTPException(status_code=400, detail="mode必须是 at_serial 或 android_adb")

    spec = _load_json_file(at_effective_atspec_path) or _load_json_file(at_spec_path)
    profile = _load_json_file(at_effective_profile_path) or _load_json_file(at_profile_path)
    efsm = _load_json_file(at_active_efsm_path) or _load_json_file(at_efsm_path)
    commands = {c.get("id"): c for c in (spec.get("commands", []) if isinstance(spec.get("commands", []), list) else []) if isinstance(c, dict) and c.get("id")}

    init_seq = ((profile.get("defaults", {}) or {}).get("init_sequence", []) if isinstance(profile.get("defaults", {}), dict) else [])
    plan = _find_transition_plan(efsm, max_steps=max(1, min(payload.max_steps, 200)))

    steps: List[Dict[str, Any]] = []
    coverage_points_total = set()
    coverage_points_hit = set()

    for t in plan:
        cov = ((t.get("coverage", {}) or {}).get("points", []) if isinstance(t.get("coverage", {}), dict) else [])
        if isinstance(cov, list):
            for c in cov:
                coverage_points_total.add(str(c))

    # execute init sequence
    for item in init_seq:
        cmd_id = item.get("cmd_id") if isinstance(item, dict) else ""
        params = item.get("params", {}) if isinstance(item, dict) else {}
        c = commands.get(cmd_id, {})
        cmd = _render_at_command(c.get("at", ""), params if isinstance(params, dict) else {})
        res = _exec_at_command(mode, cmd, payload)
        steps.append({"phase": "init", "cmd_id": cmd_id, "cmd": cmd, "result": res})

    # execute transition plan
    current_state = "S0_BOOT"
    for t in plan:
        tid = t.get("id", "")
        action = t.get("action", {}) if isinstance(t.get("action", {}), dict) else {}
        transition_ok = True
        cmd_entries: List[Dict[str, Any]] = []
        if "cmd_sequence" in action and isinstance(action.get("cmd_sequence"), list):
            for cid in action.get("cmd_sequence", []):
                c = commands.get(cid, {})
                cmd = _render_at_command(c.get("at", ""), {})
                res = _exec_at_command(mode, cmd, payload)
                cmd_entries.append({"cmd_id": cid, "cmd": cmd, "result": res})
                if not res.get("ok"):
                    transition_ok = False
        elif action.get("cmd_id"):
            cid = action.get("cmd_id")
            params = action.get("params", {}) if isinstance(action.get("params", {}), dict) else {}
            c = commands.get(cid, {})
            cmd = _render_at_command(c.get("at", ""), params)
            res = _exec_at_command(mode, cmd, payload)
            cmd_entries.append({"cmd_id": cid, "cmd": cmd, "result": res})
            transition_ok = bool(res.get("ok"))
        else:
            transition_ok = False

        if transition_ok:
            current_state = str(t.get("to", current_state))
            cov = ((t.get("coverage", {}) or {}).get("points", []) if isinstance(t.get("coverage", {}), dict) else [])
            if isinstance(cov, list):
                for c in cov:
                    coverage_points_hit.add(str(c))

        steps.append(
            {
                "phase": "transition",
                "transition_id": tid,
                "from": t.get("from"),
                "to": t.get("to"),
                "ok": transition_ok,
                "commands": cmd_entries,
            }
        )

    return {
        "mode": mode,
        "final_state": current_state,
        "coverage": {"covered": len(coverage_points_hit), "total": len(coverage_points_total), "points": sorted(list(coverage_points_hit))},
        "steps": steps,
    }


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=8010)
